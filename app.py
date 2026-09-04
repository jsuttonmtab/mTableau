import time
import functools
import sys
from pathlib import Path
import dash
from dash import Dash, html, dcc, Input, Output, State, ctx, ALL
import dash_bootstrap_components as dbc
import pandas as pd
import json
import re
from db.connection import run_query
from components.worksheet import (
    build_layout, build_table_panel, build_middle_panel,
    AVAILABLE_FIELDS, build_query, get_field_expr,
    DATE_FIELDS, DATE_FORMAT_OPTIONS, DEFAULT_MEASURE_OPTIONS
)
from pages.settings import build_settings_layout
from data.extract import (
    build_extract, get_extract_info,
    EXTRACT_PATH, USAGE_PATH, BRIDGE_PATH, STUDY_PATH, USER_GROUP_PATH
)
from components.worksheet import (
    build_layout, build_table_panel, build_middle_panel,
    AVAILABLE_FIELDS, build_query, get_field_expr,
    DATE_FIELDS, DATE_FORMAT_OPTIONS, DEFAULT_MEASURE_OPTIONS, TABLE_FIELDS
)
from utils.calculations import apply_calculation, validate_formula, FORMULA_FIELDS
from utils.query_engine import extract_available, run_extract_query, build_duckdb_query
from utils.config import load_config, save_config, get_base_dir
import threading
import duckdb

if getattr(sys, 'frozen', False):
    _base = Path(sys.executable).parent
    _assets_candidate = _base / "_internal" / "assets"
    if _assets_candidate.exists():
        _assets_folder = str(_assets_candidate)
    else:
        _assets_folder = str(_base / "assets")
else:
    _assets_folder = "assets"

app = Dash(
    __name__,
    assets_folder=_assets_folder,
    external_stylesheets=[
        dbc.themes.BOOTSTRAP,
        "https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.0/font/bootstrap-icons.css"
    ],
    suppress_callback_exceptions=True
)
app.title = "mTableau"

FMT_SHORT = {
    "year": "YEAR", "year_month": "YM", "month_name": "MON",
    "month_abbrev": "MON", "month_num": "MM", "month_num_nz": "M",
    "quarter": "QTR", "month_only": "MON", "month_only_abbrev": "MON",
    "month_only_num": "MM", "month_only_num_nz": "M"
}

def _fmt_key(shelf, field):
    return f"{shelf}|{field}"

def _get_fmt(date_formats, shelf, field):
    return date_formats.get(_fmt_key(shelf, field),
           date_formats.get(field, "none"))


# ─────────────────────────────────────────────
# Config helpers
# ─────────────────────────────────────────────

def _save_worksheets(worksheets):
    try:
        cfg = load_config()
        cfg["worksheets"] = worksheets
        save_config(cfg)
    except Exception as e:
        print(f"Could not save worksheets: {e}")


def _rename_ws_state(old_name, new_name):
    try:
        old_key = old_name.replace(" ", "_")
        new_key = new_name.replace(" ", "_")
        results_dir = get_base_dir() / "data" / "results"
        for ext in [".parquet", ".meta.json"]:
            old_file = results_dir / f"{old_key}{ext}"
            new_file = results_dir / f"{new_key}{ext}"
            if old_file.exists():
                old_file.rename(new_file)
        cfg = load_config()
        for key in ["ws_state", "ws_settings"]:
            store = cfg.get(key, {})
            if old_key in store:
                store[new_key] = store.pop(old_key)
                cfg[key] = store
        save_config(cfg)
    except Exception as e:
        print(f"Could not rename ws state: {e}")


def _duplicate_ws_state(src_name, dst_name):
    try:
        import shutil, copy
        src_key = src_name.replace(" ", "_")
        dst_key = dst_name.replace(" ", "_")
        results_dir = get_base_dir() / "data" / "results"
        for ext in [".parquet", ".meta.json"]:
            src_file = results_dir / f"{src_key}{ext}"
            dst_file = results_dir / f"{dst_key}{ext}"
            if src_file.exists():
                shutil.copy2(src_file, dst_file)
        cfg = load_config()
        for key in ["ws_state", "ws_settings"]:
            store = cfg.get(key, {})
            if src_key in store:
                store[dst_key] = copy.deepcopy(store[src_key])
                cfg[key] = store
        save_config(cfg)
    except Exception as e:
        print(f"Could not duplicate ws state: {e}")


def _save_ws_state(ws_key, rows, cols, filters, field_filters, date_formats, measure):
    try:
        cfg = load_config()
        if "ws_state" not in cfg:
            cfg["ws_state"] = {}
        cfg["ws_state"][ws_key] = {
            "rows": rows, "cols": cols, "filters": filters,
            "field_filters": field_filters, "date_formats": date_formats,
            "measure": measure,
        }
        save_config(cfg)
    except Exception as e:
        print(f"Could not save ws_state: {e}")


def clear_saved_result(ws_key):
    try:
        results_dir = get_base_dir() / "data" / "results"
        for ext in [".parquet", ".meta.json", ".summary.parquet"]:
            p = results_dir / f"{ws_key}{ext}"
            if p.exists():
                p.unlink()
    except Exception:
        pass


# ─────────────────────────────────────────────
# Filter options helper
# ─────────────────────────────────────────────\

_filter_lock = threading.Lock()
_filter_seq  = {"latest": 0}

HIGH_CARDINALITY_FIELDS = {"LONG_NAME", "EXT_STUDY_ID", "STUDYID", "USER_NAME", "USER_EMAIL"}

_STUDY_FILTER_FIELDS      = {"EXT_STUDY_ID", "LONG_NAME", "STUDYID", "STUDYYEAR"}
_USER_GROUP_FILTER_FIELDS = {"GROUP_NAME"}


def _get_filter_options_for_field(field, fmt="none"):
    if field in {"USAGE_ID", "TABRUN_TS"}:
        return []

    if field.startswith("calc_"):
        field = field[5:]

    cfg   = load_config()
    calcs = cfg.get("global_calculations", {})
    if field in calcs:
        return list(_get_options_for_field(field, fmt, calcs))

    if field in _STUDY_FILTER_FIELDS:
        parquet_src = f"read_parquet('{str(STUDY_PATH)}')"
    elif field in _USER_GROUP_FILTER_FIELDS:
        parquet_src = f"read_parquet('{str(USER_GROUP_PATH)}')"
    else:
        parquet_src = f"read_parquet('{str(USAGE_PATH)}')"

    try:
        con = duckdb.connect()
        # High cardinality fields — don't load all, require search
        if field in HIGH_CARDINALITY_FIELDS:
            count_result = con.execute(
                f"SELECT COUNT(DISTINCT {field}) as cnt FROM {parquet_src}"
            ).fetchdf()
            con.close()
            count = int(count_result["cnt"].iloc[0])
            if count > 500:
                return [{"label": "__hint__", "value": "__hint__", "disabled": True}]

        if fmt == "year":
            sql = f"""SELECT DISTINCT CAST(YEAR({field}) AS VARCHAR) as val
                      FROM {parquet_src} WHERE {field} IS NOT NULL ORDER BY val DESC"""
        elif fmt == "quarter":
            sql = f"""SELECT DISTINCT
                        CONCAT(CAST(YEAR({field}) AS VARCHAR),
                               ' Q', CAST(QUARTER({field}) AS VARCHAR)) as val
                      FROM {parquet_src} WHERE {field} IS NOT NULL
                      ORDER BY val DESC"""
        elif fmt in ("month_abbrev", "month_name", "month_num", "year_month"):
            sql = f"""SELECT DISTINCT STRFTIME({field}, '%Y-%m') as val
                      FROM {parquet_src} WHERE {field} IS NOT NULL ORDER BY val DESC"""
        elif field in {"TABRUN_MY", "ACTION_DATE"}:
            sql = f"""SELECT DISTINCT STRFTIME({field}, '%Y-%m') as val
                      FROM {parquet_src} WHERE {field} IS NOT NULL ORDER BY val DESC"""
        else:
            sql = f"""SELECT DISTINCT {field} as val
                      FROM {parquet_src} WHERE {field} IS NOT NULL ORDER BY LOWER(CAST(val AS VARCHAR))"""
        result = con.execute(sql).fetchdf()
        con.close()
        return [{"label": str(r).strip(), "value": str(r).strip()}
                for r in result["val"] if r is not None]
    except Exception as e:
        print(f"Filter options error for {field}: {e}")
        return []


@functools.lru_cache(maxsize=64)
def _cached_filter_options(field, fmt):
    return tuple(_get_filter_options_for_field(field, fmt))

def _get_options_for_field(field, fmt, calcs):
    """Get filter options for any field including global calculations."""
    if field in calcs:
        defn      = calcs[field]
        calc_type = defn.get("type")
        if calc_type == "formula":
            formula = defn.get("formula", "").upper()
            bool_funcs = {"CONTAINS", "STARTSWITH", "ENDSWITH", "ISNULL", "ISNOTNULL"}
            if any(f in formula for f in bool_funcs):
                return [{"label": "True",  "value": "True"},
                        {"label": "False", "value": "False"}]
            for date_field in {"TABRUN_MY", "TABRUN_TS", "ACTION_DATE"}:
                if date_field in formula:
                    return list(_cached_filter_options(date_field, fmt))
            for source_field in {"USER_EMAIL", "USER_NAME", "CLIENT_NAME",
                                  "GROUP_NAME", "LONG_NAME", "ACTION_TYPE"}:
                if source_field in formula:
                    return list(_cached_filter_options(source_field, fmt))
            return [{"label": "Not filterable", "value": "__none__", "disabled": True}]
        elif calc_type == "fixed_lod":
            lod_field = defn.get("field", "")
            if lod_field:
                return list(_cached_filter_options(lod_field, fmt))
            return [{"label": "Not filterable", "value": "__none__", "disabled": True}]
        elif calc_type == "aggregate":
            return [{"label": "Not filterable", "value": "__none__", "disabled": True}]
    return list(_cached_filter_options(field, fmt))

# ─────────────────────────────────────────────
# HTML Table builder
# ─────────────────────────────────────────────

def build_html_table(df, display_df=None, rows=None, max_display=1000):
    rows       = rows or []
    display_df = display_df if display_df is not None else df
    total_rows = len(df)
    truncated  = total_rows > max_display
    if truncated:
        df         = df.iloc[:max_display].copy()
        display_df = display_df.iloc[:max_display].copy()

    header = html.Thead(
        html.Tr([
            html.Th(
                [c, html.Div(className="col-resizer")],
                **{"data-col": c},
                style={
                    "position":        "relative",
                    "backgroundColor": "#2d6a4f" if c == "Grand Total"
                                       else "#1e3a5f" if c not in rows
                                       else "#0f1f3d",
                    "color":           "white",
                    "fontWeight":      "bold",
                    "fontSize":        "12px",
                    "padding":         "5px 10px",
                    "whiteSpace":      "nowrap",
                    "userSelect":      "none",
                    "minWidth":        "150px" if c in rows else "80px",
                    "cursor":          "default",
                    "borderLeft":      "2px solid #adb5bd" if c == "Grand Total"
                                       else "2px solid #adb5bd" if i > 0 and c not in rows and df.columns[i-1] in rows
                                       else "none",
                    "borderRight":     "2px solid #adb5bd" if c == "Grand Total" else "1px solid #1a3358",
                }
            )
            for i, c in enumerate(df.columns)
        ]),
        style={"position": "sticky", "top": "0", "zIndex": "2"}
    )

    body_rows = []
    prev_vals = {}
    for idx, row in df.iterrows():
        is_grand_total = str(row[df.columns[0]]) == "Grand Total"
        is_first_row   = idx == df.index[0]
        cells = []
        for col in df.columns:
            val              = row[col]
            val_str          = str(val) if val is not None else ""
            display_val      = str(display_df.loc[idx, col]) if col in display_df.columns else val_str
            is_row_field     = col in rows
            col_idx          = list(df.columns).index(col)
            is_first_measure = col_idx > 0 and col not in rows and df.columns[col_idx - 1] in rows
            is_new_val       = (
                idx != 0 and (
                    not is_row_field or
                    (val_str != "" and val_str != str(prev_vals.get(col, "")))
                )
            )
            cells.append(html.Td(
                display_val,
                style={
                    "fontSize":        "12px",
                    "padding":         "3px 10px",
                    "whiteSpace":      "normal",
                    "overflow":        "hidden",
                    "textOverflow":    "ellipsis",
                    "wordBreak":       "break-word",
                    "verticalAlign":   "top",
                    "backgroundColor": "#fff3cd" if is_grand_total else "white",
                    "fontWeight":      "bold" if is_grand_total else "normal",
                    "borderTop":       "2px solid #2d6a4f" if is_grand_total
                                       else "2px solid #adb5bd" if is_first_row
                                       else "1px solid #adb5bd" if is_new_val
                                       else "none",
                    "borderLeft":      "2px solid #adb5bd" if col == "Grand Total"
                                       else "2px solid #adb5bd" if is_first_measure
                                       else "none",
                    "borderRight":     "2px solid #adb5bd" if col == "Grand Total" else "none",
                    "borderBottom":    "none",
                    "color":           "#333" if display_val != "" else "transparent",
                }
            ))
        for col in df.columns:
            v = str(row[col]) if row[col] is not None else ""
            if v != "":
                prev_vals[col] = v
        body_rows.append(html.Tr(cells))

    truncation_notice = None
    if truncated:
        truncation_notice = html.Div(
            f"Showing first {max_display:,} of {total_rows:,} rows — Export to see all data.",
            className="text-muted text-center py-1",
            style={"fontSize": "11px", "borderTop": "1px solid #dee2e6",
                   "backgroundColor": "#f8f9fa"}
        )

    table_div = html.Div(
        html.Table(
            [header, html.Tbody(body_rows)],
            className="mtab-table",
            style={"borderCollapse": "collapse", "tableLayout": "fixed",
                   "width": "max-content"}
        ),
        style={"overflowX": "auto", "overflowY": "auto",
               "maxHeight": "calc(100vh - 220px)",
               "border": "1px solid #dee2e6", "borderRadius": "4px",
               "borderBottom": "2px solid #adb5bd", "position": "relative","userSelect": "text"}
    )
    if truncation_notice:
        return html.Div([table_div, truncation_notice])
    return table_div


# ─────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────

def apply_row_blanking(df, rows):
    display_df = df.copy()
    if rows:
        row_cols = [f for f in rows if f in display_df.columns]
        for i, field in enumerate(row_cols):
            if i == 0:
                mask = df[field] == df[field].shift(1)
                display_df.loc[mask, field] = ""
            else:
                parent_same = pd.Series([True] * len(df), index=df.index)
                for pf in row_cols[:i]:
                    parent_same = parent_same & (df[pf] == df[pf].shift(1))
                child_same = df[field] == df[field].shift(1)
                display_df.loc[parent_same & child_same, field] = ""
    return display_df


def apply_pivot(df, rows, cols, col_total="last", row_total="first", df_summary=None):
    valid_cols = [f for f in cols if f in df.columns]
    if not valid_cols:
        return df

    valid_rows = [f for f in rows if f in df.columns]

    if not valid_rows:
        try:
            grouped = df.groupby(valid_cols)["Count"].sum().reset_index()
            if len(valid_cols) == 1:
                col_field = valid_cols[0]
                values = {str(row[col_field]): row["Count"]
                          for _, row in grouped.iterrows()}
            else:
                values = {str(tuple(row[c] for c in valid_cols)): row["Count"]
                          for _, row in grouped.iterrows()}
            try:
                sorted_keys = sorted(values.keys(),
                                     key=lambda x: pd.to_datetime(x, format="%b %Y"))
            except Exception:
                sorted_keys = sorted(values.keys(), key=str)
            grand_total = int(sum(values.values()))
            if col_total == "first":
                ordered = {"Grand Total": grand_total}
                ordered.update({k: values[k] for k in sorted_keys})
            elif col_total == "hidden":
                ordered = {k: values[k] for k in sorted_keys}
            else:
                ordered = {k: values[k] for k in sorted_keys}
                ordered["Grand Total"] = grand_total
            return pd.DataFrame([ordered])
        except Exception as e:
            print(f"Pivot error (cols-only): {e}")
            return df

    try:
        pivot = df.pivot_table(
            index=valid_rows, columns=valid_cols,
            values="Count", aggfunc="sum", fill_value=0,
            margins=False, sort=False,
        )
        non_total = list(pivot.columns)
        try:
            non_total = sorted(non_total,
                               key=lambda x: pd.to_datetime(str(x), format="%b %Y"))
        except Exception:
            non_total = sorted(non_total, key=str)
        pivot         = pivot[non_total]
        pivot.columns = [str(c) for c in pivot.columns]
        non_total     = [str(c) for c in non_total]
        df_pivot      = pd.DataFrame(pivot.reset_index())
        # Sort rows case-insensitively by row fields
        if valid_rows:
            sort_keys = [df_pivot[r].astype(str).str.lower() for r in valid_rows if r in df_pivot.columns]
            if sort_keys:
                import numpy as np
                sort_df = pd.DataFrame({f"_sort_{i}": k for i, k in enumerate(sort_keys)})
                sort_order = sort_df.sort_values(list(sort_df.columns)).index
                df_pivot = df_pivot.loc[sort_order].reset_index(drop=True)

        # Grand Total column per row — sum across months (correct, no fanout)
        df_pivot["Grand Total"] = df_pivot[non_total].sum(axis=1)

        # ── Grand Total row ───────────────────────────────────────────────
        # If we have a summary df (query without fanout fields), use it
        # to get correct column totals that don't double-count.
        gt_row = {col: "" if col in valid_rows else 0
                  for col in df_pivot.columns}
        gt_row[valid_rows[0]] = "Grand Total"

        if df_summary is not None and "Count" in df_summary.columns:
            try:
                summary_valid_cols = [c for c in valid_cols if c in df_summary.columns]
                if summary_valid_cols:
                    summary_pivot = df_summary.pivot_table(
                        columns=summary_valid_cols,
                        values="Count",
                        aggfunc="sum",
                        fill_value=0,
                    )
                    # summary_pivot columns are the col values (e.g. month names)
                    sp_cols = {str(c): v for c, v in
                               summary_pivot.iloc[0].items()}
                    for c in non_total:
                        gt_row[c] = int(sp_cols.get(c, 0))
                else:
                    # No col fields in summary — use total Count
                    gt_row_val = int(df_summary["Count"].sum())
                    for c in non_total:
                        gt_row[c] = int(df_pivot[c].sum())
            except Exception as e:
                print(f"Summary pivot error: {e}")
                for c in non_total:
                    gt_row[c] = int(df_pivot[c].sum())
        else:
            # No fanout — column sums are correct
            for c in non_total:
                gt_row[c] = int(df_pivot[c].sum())

        gt_row["Grand Total"] = sum(gt_row[c] for c in non_total)

        gt_df    = pd.DataFrame([gt_row])
        df_pivot = pd.concat([df_pivot, gt_df], ignore_index=True)

        if col_total == "first":
            col_order = valid_rows + ["Grand Total"] + non_total
        elif col_total == "hidden":
            col_order = valid_rows + non_total
        else:
            col_order = valid_rows + non_total + ["Grand Total"]

        df_pivot = df_pivot[[c for c in col_order if c in df_pivot.columns]]

        grand_rows = df_pivot[df_pivot[valid_rows[0]] == "Grand Total"]
        other_rows = df_pivot[df_pivot[valid_rows[0]] != "Grand Total"]
        if row_total == "hidden":
            df = other_rows.reset_index(drop=True)
        elif row_total == "last":
            df = pd.concat([other_rows, grand_rows]).reset_index(drop=True)
        else:
            df = pd.concat([grand_rows, other_rows]).reset_index(drop=True)

    except Exception as e:
        print(f"Pivot error: {e}")
    return df


def build_duck_where(field_filters, date_formats):
    from utils.query_engine import FIELD_REGISTRY
    from utils.config import load_config as _load_config
    duck_where = []
    cfg   = _load_config()
    calcs = cfg.get("global_calculations", {})

    for shelf_field, saved in (field_filters or {}).items():
        if not saved:
            continue

        # Support both old list format and new dict format
        if isinstance(saved, dict):
            values  = saved.get("values", [])
            exclude = saved.get("exclude", False)
        else:
            values  = saved
            exclude = False

        if not values:
            continue

        if "|" in shelf_field:
            shelf, field = shelf_field.split("|", 1)
        else:
            shelf, field = "filters", shelf_field

        clean_field = field[5:] if field.startswith("calc_") else field

        # Formula calculation filter
        if clean_field in calcs:
            defn = calcs[clean_field]
            if defn.get("type") == "formula":
                formula = defn.get("formula", "")
                for fname, finfo in FIELD_REGISTRY.items():
                    formula = formula.replace(fname, finfo["col"])
                vals      = ", ".join([f"'{v.strip().lower().replace(chr(39), chr(39)+chr(39))}'" for v in values])
                not_kw    = "NOT " if exclude else ""
                duck_where.append(
                    f"LOWER(CAST(({formula}) AS VARCHAR)) {not_kw}IN ({vals})"
                )
                continue

        col = FIELD_REGISTRY.get(clean_field, {}).get("col", clean_field)

        # Skip unknown fields
        if clean_field not in FIELD_REGISTRY and clean_field not in calcs:
            continue

        if clean_field in {"TABRUN_MY", "ACTION_DATE", "TABRUN_TS"}:
            clauses = []
            for v in values:
                v = v.strip()
                # Handle quarter format: "Q1 2025" or "2025 Q1"
                q_match = re.match(r'^Q(\d)\s+(\d{4})$', v, re.IGNORECASE)
                if not q_match:
                    q_match = re.match(r'^(\d{4})\s+Q(\d)$', v, re.IGNORECASE)
                    if q_match:
                        yr, qtr = q_match.group(1), q_match.group(2)
                    else:
                        yr, qtr = None, None
                else:
                    qtr, yr = q_match.group(1), q_match.group(2)

                if yr and qtr:
                    qtr = int(qtr)
                    start_month = (qtr - 1) * 3 + 1
                    end_month = start_month + 2
                    if exclude:
                        clauses.append(
                            f"NOT (YEAR({col}) = {yr} AND QUARTER({col}) = {qtr})"
                        )
                    else:
                        clauses.append(
                            f"(YEAR({col}) = {yr} AND QUARTER({col}) = {qtr})"
                        )
                else:
                    # Regular LIKE for year, month, etc.
                    if exclude:
                        clauses.append(f"{col}::VARCHAR NOT LIKE '%{v}%'")
                    else:
                        clauses.append(f"{col}::VARCHAR LIKE '%{v}%'")

            if exclude:
                duck_where.append(f"({' AND '.join(clauses)})")
            else:
                duck_where.append(f"({' OR '.join(clauses)})")
        else:
            vals   = ", ".join([f"'{v.strip().replace(chr(39), chr(39)+chr(39))}'" for v in values])
            not_kw = "NOT " if exclude else ""
            duck_where.append(f"{col} {not_kw}IN ({vals})")

    return duck_where

def _build_filter_summary(values, exclude=False):
    if not values:
        return "None"
    prefix = "NOT: " if exclude else ""
    if len(values) <= 3:
        return prefix + ", ".join(str(v) for v in values)
    return prefix + f"{', '.join(str(v) for v in values[:3])} +{len(values)-3} more"

def build_ws_wrapper(w, display="none"):
    return html.Div(
        build_layout(w.replace(" ", "_")),
        id={"type": "ws-wrapper", "index": w.replace(" ", "_")},
        style={"height": "100%", "display": display}
    )


# ─────────────────────────────────────────────
# Shelf render helpers
# ─────────────────────────────────────────────

def _render_rows_items(rows, field_filters, date_formats, worksheet_id):
    if not rows:
        return html.Span("Drop fields here", className="text-muted fst-italic",
                         style={"fontSize": "11px"})
    items = []
    for f in rows:
        fmt         = _get_fmt(date_formats, "rows", f)
        label       = f if fmt == "none" else f"{f} [{FMT_SHORT.get(fmt, fmt)}]"
        is_filtered = bool(field_filters.get(f) or field_filters.get(f"filters|{f}"))
        items.append(html.Span([
            html.Span(
                [label] + ([html.I(className="bi bi-funnel-fill ms-1",
                                  style={"fontSize": "9px", "color": "yellow",
                                         "pointerEvents": "none"})]
                            if is_filtered else []),
                style={"fontSize": "11px", "padding": "2px 8px",
                       "backgroundColor": "#0a58ca", "color": "white",
                       "borderRadius": "4px 0 0 4px",
                       "display": "inline-flex", "alignItems": "center",
                       "cursor": "default", "pointerEvents": "none"}
            ),
            dbc.Button("×",
                id={"type": "remove-rows-btn", "index": f"{worksheet_id}|{f}"},
                color="primary", size="sm",
                style={"fontSize": "13px", "padding": "2px 5px",
                       "borderRadius": "0 4px 4px 0"}),
        ], className="me-1 d-inline-flex draggable-badge", draggable="true",
           **{"data-field": f, "data-source-shelf": "rows",
              "data-worksheet": worksheet_id, "data-fmt": fmt}))
    return items


def _render_cols_items(cols, field_filters, date_formats, worksheet_id):
    if not cols:
        return html.Span("Drop fields here", className="text-muted fst-italic",
                         style={"fontSize": "11px"})
    items = []
    for f in cols:
        fmt         = _get_fmt(date_formats, "cols", f)
        label       = f if fmt == "none" else f"{f} [{FMT_SHORT.get(fmt, fmt)}]"
        is_filtered = bool(field_filters.get(f) or field_filters.get(f"filters|{f}"))
        items.append(html.Span([
            html.Span(
                [label] + ([html.I(className="bi bi-funnel-fill ms-1",
                                  style={"fontSize": "9px", "color": "yellow",
                                         "pointerEvents": "none"})]
                            if is_filtered else []),
                style={"fontSize": "11px", "padding": "2px 8px",
                       "backgroundColor": "#157347", "color": "white",
                       "borderRadius": "4px 0 0 4px",
                       "display": "inline-flex", "alignItems": "center",
                       "cursor": "default", "pointerEvents": "none"}
            ),
            dbc.Button("×",
                id={"type": "remove-cols-btn", "index": f"{worksheet_id}|{f}"},
                color="success", size="sm",
                style={"fontSize": "13px", "padding": "2px 5px",
                       "borderRadius": "0 4px 4px 0"}),
        ], className="me-1 d-inline-flex draggable-badge", draggable="true",
           **{"data-field": f, "data-source-shelf": "cols",
              "data-worksheet": worksheet_id, "data-fmt": fmt}))
    return items


# ─────────────────────────────────────────────
# Layout
# ─────────────────────────────────────────────

_initial_worksheets = load_config().get("worksheets", ["Worksheet 1"])

app.layout = html.Div([
    dcc.Store(id="worksheet-store",        data=_initial_worksheets, storage_type="memory"),
    dcc.Store(id="global-calculations",    data=load_config().get("global_calculations", {})),
    dcc.Store(id="extract-running",        data=False),
    dcc.Store(id="rename-worksheet-input", data=""),
    dcc.Store(id="drop-payload",           data=""),
    dcc.Store(id="dupe-payload",           data=""),
    dcc.Store(id="ws-settings-payload",    data=""),
    dcc.Store(id="ws-settings-store",      data={}),
    dcc.Store(id="editing-calc-name",      data=""),
    dcc.Store(id="extract-trigger",        data=0),
    dcc.Store(id="badge-context-payload",  data=""),
    dcc.Store(id="ws-last-run-state",      data={}, storage_type="memory"),
    dcc.Store(id="query-cancel-signal",    data=False),
    dcc.Store(id="delete-ws-payload",      data=""),
    dcc.Store(id="filter-pending-search",  data=""),
    dcc.Store(id="delete-ws-pending", data=""),
    dcc.Download(id="download-data"),
    dcc.Download(id="download-crosstab"),

    html.Button(id="rename-trigger-btn",        style={"display": "none"}),
    html.Button(id="drop-trigger-btn",          style={"display": "none"}),
    html.Button(id="dupe-trigger-btn",          style={"display": "none"}),
    html.Button(id="ws-settings-trigger-btn",   style={"display": "none"}),
    html.Button(id="badge-context-trigger-btn", style={"display": "none"}),
    html.Button(id="delete-ws-trigger-btn",     style={"display": "none"}),

    html.Div([
        html.Button(id="refresh-extract-btn", n_clicks=0),
        html.Button(id="cancel-extract-btn",  n_clicks=0),
    ], style={"display": "none"}),

    # ── Modals ────────────────────────────────
    dbc.Modal([
        dbc.ModalHeader(id="perm-field-filter-header", style={"padding": "8px 16px"}),
        dbc.ModalBody([
            html.Div([
                html.Span("Selected: ", className="fw-bold", style={"fontSize": "11px"}),
                html.Span(id="perm-field-filter-summary",
                         className="text-primary",
                         style={"fontSize": "11px", "cursor": "pointer",
                                "textDecoration": "underline"})
            ], id="filter-summary-clickable",
               className="mb-2 p-2 border rounded",
               style={"backgroundColor": "#f0f4ff", "cursor": "pointer"}),
            html.Div([
                dbc.Input(id="perm-filter-search", placeholder="Search...",
                         type="text",
                         className="form-control form-control-sm",
                         style={"fontSize": "12px", "flex": "1"}),
                dbc.Button("Search", id="perm-filter-search-btn", size="sm",
                          color="primary",
                          style={"fontSize": "11px", "padding": "1px 10px",
                                 "marginLeft": "6px", "whiteSpace": "nowrap"}),
                dbc.Button("Clear", id="perm-filter-clear-search-btn", size="sm",
                          color="outline-secondary",
                          style={"fontSize": "11px", "padding": "1px 8px",
                                 "marginLeft": "4px", "whiteSpace": "nowrap"}),
            ], className="d-flex align-items-center mb-2"),
            dcc.Store(id="perm-filter-sort", data={}),
            html.Div([
                dbc.Label("Date Format:", size="sm", className="fw-bold me-2",
                         style={"fontSize": "11px", "whiteSpace": "nowrap"}),
                dbc.Select(id="perm-filter-date-fmt",
                          options=[{"label": "None", "value": "none"}],
                          value="none", size="sm",
                          disabled=True,
                          style={"fontSize": "11px", "maxWidth": "200px"}),
            ], id="perm-filter-fmt-row",
               className="d-flex align-items-center mb-2"),
            html.Div(id="filter-search-status",
                    style={"fontSize": "10px", "color": "#888",
                           "marginBottom": "4px", "minHeight": "14px"}),
            html.Div(
                id="filter-loading-msg",
                children=[
                    html.Div("⏳", style={"fontSize": "28px", "marginBottom": "8px"}),
                    html.Div("Loading options, please wait...",
                             style={"fontSize": "13px", "fontWeight": "bold"}),
                ],
                className="filter-loading-pulse",
                style={"display": "none", "textAlign": "center",
                       "padding": "30px", "color": "#0f1f3d"}
            ),
            html.Div([
                dbc.Button("Select All", id="filter-select-all", size="sm", color="link",
                          style={"fontSize": "11px", "padding": "0 8px 4px 0"}),
                dbc.Button("Select None", id="filter-select-none", size="sm", color="link",
                          style={"fontSize": "11px", "padding": "0 0 4px 0",
                                 "color": "#dc3545"}),
                dbc.Button("Select Range", id="filter-select-range", size="sm", color="link",
                          style={"fontSize": "11px", "padding": "0 0 4px 8px",
                                 "color": "#0d6efd"}),
                dbc.Button("A→Z", id="perm-filter-sort-btn", size="sm",
                          color="outline-secondary",
                          className="ms-auto",
                          style={"fontSize": "10px", "padding": "1px 8px",
                                 "whiteSpace": "nowrap"}),
            ], className="d-flex align-items-center mb-1"),
            html.Div([
                dbc.Checklist(id="perm-field-filter-checklist", options=[], value=[],
                             className="small", style={"fontSize": "12px"})
            ], style={"overflowY": "auto", "maxHeight": "300px",
                     "border": "1px solid #dee2e6", "borderRadius": "4px",
                     "padding": "8px"}),
        ], style={"padding": "8px 16px"}),
        dbc.ModalFooter([
            html.Div([
                dbc.Checkbox(id="perm-filter-exclude",
                            label="Exclude selected",
                            value=False,
                            style={"fontSize": "12px"}),
                dbc.Checkbox(id="perm-filter-cascade",
                            label="Apply existing filters",
                            value=False,
                            style={"fontSize": "12px"}),
            ], className="me-auto d-flex flex-column gap-1"),
            dbc.Button("Clear Filter", id="perm-field-filter-clear",
                      color="secondary", size="sm", className="me-2"),
            dbc.Button("Apply", id="perm-field-filter-apply",
                      color="primary", size="sm"),
        ], style={"padding": "8px 16px", "display": "flex",
                  "alignItems": "center"}),
        dcc.Store(id="perm-field-context", data={}),
    ], id="perm-field-filter-modal", is_open=False, size="lg"),

    dbc.Modal([
        dbc.ModalHeader("Select Range", style={"padding": "8px 16px"}),
        dbc.ModalBody([
            html.Div([
                dbc.Label("From:", size="sm", className="fw-bold",
                         style={"fontSize": "12px"}),
                dbc.Select(id="range-from-select", options=[], size="sm",
                          style={"fontSize": "11px"}),
            ], className="mb-2"),
            html.Div([
                dbc.Label("To:", size="sm", className="fw-bold",
                         style={"fontSize": "12px"}),
                dbc.Select(id="range-to-select", options=[], size="sm",
                          style={"fontSize": "11px"}),
            ], className="mb-2"),
        ], style={"padding": "8px 16px"}),
        dbc.ModalFooter([
            dbc.Button("Cancel", id="range-cancel-btn",
                      color="secondary", size="sm", className="me-2"),
            dbc.Button("Select", id="range-apply-btn",
                      color="primary", size="sm"),
        ], style={"padding": "8px 16px"}),
    ], id="range-select-modal", is_open=False, size="sm", centered=True),

    dbc.Modal([
        dbc.ModalHeader("Selected Values", style={"padding": "8px 16px"}),
        dbc.ModalBody(id="selected-values-body",
                     style={"padding": "8px 16px", "maxHeight": "400px",
                            "overflowY": "auto"}),
        dbc.ModalFooter(
            dbc.Button("Close", id="selected-values-close",
                      color="secondary", size="sm"),
            style={"padding": "8px 16px"}
        ),
    ], id="selected-values-modal", is_open=False, size="md", centered=True),

    dbc.Modal([
        dbc.ModalHeader("➕ Add Calculation"),
        dbc.ModalBody([
            dbc.Label("Name", size="sm", className="fw-bold"),
            dbc.Input(id="calc-name-input", placeholder="e.g. Distinct Users",
                     size="sm", className="mb-3"),
            dbc.Label("Type", size="sm", className="fw-bold"),
            dbc.RadioItems(
                id="calc-type-select",
                options=[
                    {"label": "Aggregate (COUNTD, COUNT, SUM, AVG)", "value": "aggregate"},
                    {"label": "Formula (YEAR(TABRUN_MY), UPPER(CLIENT_NAME)…)", "value": "formula"},
                ],
                value="aggregate", className="mb-3", style={"fontSize": "13px"}
            ),
            html.Div(id="calc-aggregate-section", children=[
                dbc.Row([
                    dbc.Col([
                        dbc.Label("Function", size="sm", className="fw-bold"),
                        dbc.Select(id="calc-agg-func", options=[
                            {"label": "COUNTD (distinct count)", "value": "COUNTD"},
                            {"label": "COUNT (total rows)",       "value": "COUNT"},
                            {"label": "SUM",                      "value": "SUM"},
                            {"label": "AVG",                      "value": "AVG"},
                        ], value="COUNTD", size="sm"),
                    ], width=5),
                    dbc.Col([
                        dbc.Label("Field", size="sm", className="fw-bold"),
                        dbc.Select(id="calc-agg-field", options=[
                            {"label": f, "value": f} for f in
                            ["USAGE_ID", "USER_NAME", "CLIENT_NAME",
                             "EXT_STUDY_ID", "GROUP_NAME"]
                        ], value="USAGE_ID", size="sm"),
                    ], width=7),
                ]),
            ]),
            html.Div(id="calc-formula-section", style={"display": "none"}, children=[
                dbc.Label("Formula", size="sm", className="fw-bold"),
                dbc.Textarea(id="calc-formula-input",
                            placeholder="e.g. YEAR(TABRUN_MY)", rows=3,
                            className="mb-2",
                            style={"fontSize": "13px", "fontFamily": "monospace"}),
            ]),
            html.Div(id="calc-lod-section", style={"display": "none"}, children=[
                dbc.Row([
                    dbc.Col([
                        dbc.Label("Fixed By", size="sm", className="fw-bold"),
                        dbc.Select(id="calc-lod-dim", options=[
                            {"label": f, "value": f} for f in
                            ["EXT_STUDY_ID", "USER_NAME", "CLIENT_NAME",
                             "GROUP_NAME", "LONG_NAME", "STUDYID"]
                        ], value="EXT_STUDY_ID", size="sm"),
                    ], width=6),
                    dbc.Col([
                        dbc.Label("Agg", size="sm", className="fw-bold"),
                        dbc.Select(id="calc-lod-agg", options=[
                            {"label": "MAX", "value": "MAX"},
                            {"label": "MIN", "value": "MIN"},
                            {"label": "AVG", "value": "AVG"},
                            {"label": "SUM", "value": "SUM"},
                            {"label": "COUNT", "value": "COUNT"},
                        ], value="MAX", size="sm"),
                    ], width=3),
                    dbc.Col([
                        dbc.Label("Of Field", size="sm", className="fw-bold"),
                        dbc.Select(id="calc-lod-field", options=[
                            {"label": f, "value": f} for f in
                            ["TABRUN_TS", "TABRUN_MY", "USAGE_ID"]
                        ], value="TABRUN_TS", size="sm"),
                    ], width=3),
                ], className="mb-2"),
                html.Div(id="calc-lod-preview", className="mt-2 p-2 rounded",
                        style={"fontSize": "11px", "color": "#555",
                               "backgroundColor": "#f8f9fa",
                               "fontFamily": "monospace", "minHeight": "24px"}),
            ]),
            html.Div(id="calc-validation-msg", className="mt-2",
                    style={"fontSize": "12px"}),
        ]),
        dbc.ModalFooter([
            dbc.Button("Cancel", id="calc-cancel-btn", color="secondary",
                      size="sm", className="me-2"),
            dbc.Button("Save", id="calc-save-btn", color="primary", size="sm"),
        ])
    ], id="calc-modal", is_open=False, size="md"),

    dbc.Modal([
        dbc.ModalHeader("New Worksheet"),
        dbc.ModalBody([
            dbc.Label("Worksheet Name:"),
            dbc.Input(id="new-worksheet-name",
                     placeholder="e.g. Client Usage by Month",
                     type="text", autoComplete="new-password")
        ]),
        dbc.ModalFooter([
            dbc.Button("Cancel", id="cancel-worksheet-btn",
                      color="secondary", className="me-2"),
            dbc.Button("Create", id="confirm-worksheet-btn", color="primary"),
        ])
    ], id="new-worksheet-modal", is_open=False),

    dbc.Modal([
        dbc.ModalHeader("⚙️ Settings"),
        dbc.ModalBody(id="settings-modal-body", style={"padding": "0"}),
    ], id="settings-modal", is_open=False, size="xl", scrollable=True),

    dbc.Modal([
        dbc.ModalBody([
            html.Div([
                dbc.Spinner(color="primary", size="lg"),
                html.Div("Running query...", className="mt-3 fw-bold",
                        style={"fontSize": "16px"}),
                html.Div("Please wait", className="text-muted",
                        style={"fontSize": "13px"}),
                dbc.Button("Cancel", id="cancel-query-btn",
                          color="danger", outline=True, size="sm",
                          className="mt-3"),
            ], className="d-flex flex-column align-items-center py-3")
        ])
    ], id="query-running-modal", is_open=False, centered=True,
       size="sm", backdrop="static", keyboard=False),
       
   dbc.Modal([
        dbc.ModalBody(
            id="delete-ws-confirm-body",
            style={"fontSize": "14px", "padding": "20px"}
        ),
        dbc.ModalFooter([
            dbc.Button("Cancel", id="delete-ws-cancel-btn",
                      color="secondary", size="sm", className="me-2"),
            dbc.Button("Delete", id="delete-ws-confirm-btn",
                      color="danger", size="sm"),
        ], style={"padding": "8px 16px"}),
    ], id="delete-ws-confirm-modal", is_open=False, centered=True, size="sm"),

    dbc.Modal([
        dbc.ModalHeader(id="ws-settings-header"),
        dbc.ModalBody([
            dbc.Row([
                dbc.Col([
                    dbc.Label("Column Grand Total", className="fw-bold", size="sm"),
                    dbc.Select(id="ws-col-grand-total",
                        options=[
                            {"label": "Last column (default)", "value": "last"},
                            {"label": "First column",          "value": "first"},
                            {"label": "Hidden",                "value": "hidden"},
                        ], value="last", size="sm"),
                ], width=6),
                dbc.Col([
                    dbc.Label("Row Grand Total", className="fw-bold", size="sm"),
                    dbc.Select(id="ws-row-grand-total",
                        options=[
                            {"label": "First row (default)", "value": "first"},
                            {"label": "Last row",            "value": "last"},
                            {"label": "Hidden",              "value": "hidden"},
                        ], value="first", size="sm"),
                ], width=6),
            ]),
        ], className="pt-3"),
        dbc.ModalFooter([
            dbc.Button("Cancel", id="ws-settings-cancel-btn",
                      color="secondary", size="sm", className="me-2"),
            dbc.Button("Apply", id="ws-settings-apply-btn",
                      color="primary", size="sm"),
        ]),
    ], id="ws-settings-modal", is_open=False, size="md", centered=True),

    # ── Main visible layout ───────────────────
    dbc.Navbar(
        dbc.Container([
            html.Img(src="/assets/mtableauLogo2.png", height="40px", className="me-2"),
            dbc.NavbarBrand("", className="fw-bold fs-5 text-white"),
            dbc.Button(html.I(className="bi bi-gear-fill"),
                      id="settings-btn", color="light", size="sm",
                      className="ms-auto", title="Settings"),
        ], fluid=True),
        color="dark",
        style={"backgroundColor": "#0f1f3d"}, className="mb-0 py-1"
    ),

    html.Div(id="extract-progress-persistent", style={"display": "none"}, children=[
        html.Div([
            html.Span(id="extract-progress-persistent-msg",
                     style={"fontSize": "12px", "marginRight": "10px"}),
            dbc.Progress(id="extract-progress-persistent-bar", value=0,
                        striped=True, animated=True,
                        style={"height": "6px", "flex": "1"}),
        ], className="d-flex align-items-center px-3 py-1",
           style={"backgroundColor": "#e8f4ff", "borderBottom": "1px solid #bee3f8"})
    ]),

    html.Div(id="extract-clock", style={"display": "none"}, children=[
        html.Div(id="extract-clock-display", className="text-center py-1",
                style={"fontSize": "11px", "color": "#666",
                       "backgroundColor": "#f0f4ff",
                       "borderBottom": "1px solid #dee2e6"})
    ]),

    dcc.Interval(id="extract-clock-interval", interval=30000, disabled=True),
    dcc.Interval(id="extract-interval",       interval=2000,  disabled=True),
    html.Div(id="extract-progress",               style={"display": "none"}),
    html.Div(id="extract-progress-bar-container", style={"display": "none"}, children=[
        dbc.Progress(id="extract-progress-bar", value=0, striped=True, animated=True,
                    style={"height": "20px"})
    ]),

    html.Div(id="worksheet-content",
            style={"flex": "1", "overflow": "hidden", "minHeight": "0"}),

    html.Div([
        html.Div(style={"width": "175px", "flexShrink": "0"}),
        html.Div([
            dcc.Tabs(
                id="worksheet-tabs",
                value=_initial_worksheets[0],
                children=[dcc.Tab(label=w, value=w) for w in _initial_worksheets],
                className="custom-tabs",
                style={"height": "34px"},
            ),
        ], style={"flex": "1", "overflow": "hidden", "position": "relative"}),
        dbc.Button("+ Sheet", id="add-worksheet-btn",
                  color="outline-secondary", size="sm",
                  style={"fontSize": "11px", "margin": "3px 6px",
                         "flexShrink": "0", "height": "28px",
                         "alignSelf": "center"}),
    ], className="d-flex align-items-center border-top",
       style={"backgroundColor": "#f8f8f8", "flexShrink": "0", "height": "38px"}),

], style={"display": "flex", "flexDirection": "column",
          "height": "100vh", "overflow": "hidden"})

@app.callback(
    Output("perm-filter-sort",     "data"),
    Output("perm-filter-sort-btn", "children"),
    Input("perm-filter-sort-btn",  "n_clicks"),
    State("perm-filter-sort",      "data"),
    State("perm-field-context",    "data"),
    prevent_initial_call=True
)
def toggle_filter_sort(n_clicks, sort_data, context):
    sort_data = sort_data if isinstance(sort_data, dict) else {}
    field = (context or {}).get("field", "")
    current = sort_data.get(field, "asc")
    if current == "asc":
        sort_data[field] = "desc"
        return sort_data, "Z→A"
    sort_data[field] = "asc"
    return sort_data, "A→Z"
    
@app.callback(
    Output("filter-loading-msg", "style"),
    Output("filter-loading-msg", "children"),
    Input("perm-field-filter-checklist", "options"),
    Input("perm-field-filter-modal",     "is_open"),
    State("perm-filter-search",          "value"),
    State("perm-field-context",          "data"),
    prevent_initial_call=True
)
def toggle_loading_msg(options, is_open, search, context):
    hidden     = {"display": "none"}
    spinner    = {"display": "block", "textAlign": "center",
                  "padding": "30px", "color": "#0f1f3d",
                  "fontSize": "13px", "fontWeight": "bold"}
    hint_style = {"display": "block", "textAlign": "center",
                  "padding": "20px", "color": "#666", "fontSize": "13px"}

    spinner_children = [
        html.Div("⏳", style={"fontSize": "28px", "marginBottom": "8px"}),
        html.Div("Loading options, please wait...",
                 style={"fontSize": "13px", "fontWeight": "bold"}),
    ]
    hint_children = [
        html.Div("🔍", style={"fontSize": "24px", "marginBottom": "8px"}),
        html.Div("Please type at least 3 characters to begin search",
                 style={"fontSize": "13px", "color": "#555"}),
    ]

    if not is_open:
        return hidden, spinner_children

    field      = (context or {}).get("field", "")
    search_len = len((search or "").strip())

    if field in HIGH_CARDINALITY_FIELDS:
        if search_len < 3:
            return hint_style, hint_children
        if not options or options == [] or \
           (len(options) == 1 and options[0].get("value") == "__cleared__"):
            return spinner, spinner_children
        return hidden, spinner_children

    is_loading = (not options or
                  (len(options) == 1 and
                   options[0].get("value") in ("__loading__", "__searching__")))
    if is_loading:
        return spinner, spinner_children
    return hidden, spinner_children


@app.callback(
    Output("perm-filter-search", "value", allow_duplicate=True),
    Input("perm-filter-clear-search-btn", "n_clicks"),
    prevent_initial_call=True
)
def clear_search_btn(n_clicks):
    if not n_clicks:
        raise dash.exceptions.PreventUpdate
    return ""

# ─────────────────────────────────────────────
# Clientside callbacks
# ─────────────────────────────────────────────

app.clientside_callback(
    """function(n) {
        if (!n) return '';
        var p = window._dashDropPayload || '';
        if (!p) return '';
        try {
            var d = JSON.parse(p);
            d._t = Date.now();
            return JSON.stringify(d);
        } catch(e) { return p + '|' + Date.now(); }
    }""",
    Output("drop-payload", "data"),
    Input("drop-trigger-btn", "n_clicks"),
    prevent_initial_call=True
)
app.clientside_callback(
    "function(n) { return window._dashRenamePayload || ''; }",
    Output("rename-worksheet-input", "data"),
    Input("rename-trigger-btn", "n_clicks"),
    prevent_initial_call=True
)
app.clientside_callback(
    "function(n) { return window._dashDupePayload || ''; }",
    Output("dupe-payload", "data"),
    Input("dupe-trigger-btn", "n_clicks"),
    prevent_initial_call=True
)
app.clientside_callback(
    "function(n) { return window._dashWsSettingsPayload || ''; }",
    Output("ws-settings-payload", "data"),
    Input("ws-settings-trigger-btn", "n_clicks"),
    prevent_initial_call=True
)
app.clientside_callback(
    "function(n) { return window._dashBadgeContextPayload || ''; }",
    Output("badge-context-payload", "data"),
    Input("badge-context-trigger-btn", "n_clicks"),
    prevent_initial_call=True
)
app.clientside_callback(
    """function(n_clicks) {
        if (n_clicks && n_clicks.some(n => n)) return true;
        return window.dash_clientside.no_update;
    }""",
    Output("query-running-modal", "is_open"),
    Input({"type": "run-query-btn", "index": ALL}, "n_clicks"),
    prevent_initial_call=True
)
app.clientside_callback(
    "function(n) { return window._dashDeletePayload || ''; }",
    Output("delete-ws-payload", "data"),
    Input("delete-ws-trigger-btn", "n_clicks"),
    prevent_initial_call=True
)


# ─────────────────────────────────────────────
# Filter modal helpers
# ─────────────────────────────────────────────

@app.callback(
    Output("query-cancel-signal",  "data"),
    Output("query-running-modal",  "is_open", allow_duplicate=True),
    Input("cancel-query-btn",      "n_clicks"),
    prevent_initial_call=True
)
def cancel_query(n_clicks):
    if n_clicks:
        return True, False
    raise dash.exceptions.PreventUpdate



@app.callback(
    Output("perm-field-filter-checklist", "value", allow_duplicate=True),
    Input("filter-select-all",  "n_clicks"),
    Input("filter-select-none", "n_clicks"),
    State("perm-field-filter-checklist", "options"),
    prevent_initial_call=True
)
def select_all_none(select_all, select_none, options):
    triggered = ctx.triggered_id
    if not options:
        raise dash.exceptions.PreventUpdate
    if triggered == "filter-select-all":
        return [o["value"] for o in options
                if o.get("value") not in ("__loading__", "__error__",
                                          "__hint__", "__none__")
                and not o.get("disabled")]
    elif triggered == "filter-select-none":
        return []
    raise dash.exceptions.PreventUpdate

@app.callback(
    Output("range-select-modal", "is_open"),
    Output("range-from-select",  "options"),
    Output("range-to-select",    "options"),
    Input("filter-select-range", "n_clicks"),
    Input("range-cancel-btn",    "n_clicks"),
    Input("range-apply-btn",     "n_clicks"),
    State("perm-field-filter-checklist", "options"),
    prevent_initial_call=True
)
def toggle_range_modal(select_clicks, cancel_clicks, apply_clicks, options):
    triggered = ctx.triggered_id
    if triggered in ("range-cancel-btn", "range-apply-btn"):
        return False, dash.no_update, dash.no_update
    if triggered == "filter-select-range":
        if not options:
            raise dash.exceptions.PreventUpdate
        valid = [o for o in options
                 if o.get("value") not in ("__loading__", "__error__",
                                           "__hint__", "__none__")
                 and not o.get("disabled")]
        select_options = [{"label": o["label"], "value": o["value"]} for o in valid]
        return True, select_options, select_options
    raise dash.exceptions.PreventUpdate


@app.callback(
    Output("perm-field-filter-checklist", "value", allow_duplicate=True),
    Output("range-select-modal",          "is_open", allow_duplicate=True),
    Input("range-apply-btn", "n_clicks"),
    State("range-from-select",               "value"),
    State("range-to-select",                 "value"),
    State("perm-field-filter-checklist",     "options"),
    State("perm-field-filter-checklist",     "value"),
    prevent_initial_call=True
)
def apply_range_selection(n_clicks, from_val, to_val, options, current_values):
    if not n_clicks or not from_val or not to_val:
        raise dash.exceptions.PreventUpdate
    valid = [o["value"] for o in options
             if o.get("value") not in ("__loading__", "__error__",
                                       "__hint__", "__none__")
             and not o.get("disabled")]
    try:
        from_idx = valid.index(from_val)
        to_idx   = valid.index(to_val)
    except ValueError:
        raise dash.exceptions.PreventUpdate
    start = min(from_idx, to_idx)
    end   = max(from_idx, to_idx)
    range_values = valid[start:end + 1]
    # Add range to current selection
    new_values = list(set((current_values or []) + range_values))
    return new_values, False

@app.callback(
    Output("selected-values-modal", "is_open"),
    Output("selected-values-body",  "children"),
    Input("filter-summary-clickable", "n_clicks"),
    Input("selected-values-close",    "n_clicks"),
    State("perm-field-filter-checklist", "value"),
    State("perm-field-context",          "data"),
    prevent_initial_call=True
)
def toggle_selected_values(open_clicks, close_clicks, current_values, context):
    triggered = ctx.triggered_id
    if triggered == "selected-values-close":
        return False, dash.no_update
    if not open_clicks:
        raise dash.exceptions.PreventUpdate
    if not current_values:
        return True, html.P("No values selected.", className="text-muted fst-italic",
                            style={"fontSize": "12px"})
    field = (context or {}).get("field", "")
    items = [
        html.Div([
            html.Span(str(v), style={"fontSize": "12px", "flex": "1"}),
            dbc.Button("×", id={"type": "remove-selected-value", "index": str(v)},
                      color="danger", size="sm", outline=True,
                      style={"fontSize": "11px", "padding": "1px 6px"}),
        ], className="d-flex align-items-center justify-content-between mb-1 px-2 py-1",
           style={"backgroundColor": "#f8f9fa", "borderRadius": "4px",
                  "border": "1px solid #dee2e6"})
        for v in current_values
    ]
    return True, html.Div([
        html.Div(f"{len(current_values)} value{'s' if len(current_values) != 1 else ''} selected for {field}",
                className="text-muted mb-2", style={"fontSize": "11px"}),
        html.Div(items)
    ])


@app.callback(
    Output("perm-field-filter-checklist", "value", allow_duplicate=True),
    Output("selected-values-modal",        "is_open", allow_duplicate=True),
    Output("selected-values-body",         "children", allow_duplicate=True),
    Input({"type": "remove-selected-value", "index": ALL}, "n_clicks"),
    State("perm-field-filter-checklist", "value"),
    State("perm-field-context",          "data"),
    prevent_initial_call=True
)
def remove_selected_value(n_clicks, current_values, context):
    if not any(n for n in n_clicks if n):
        raise dash.exceptions.PreventUpdate
    triggered = ctx.triggered_id
    if triggered is None:
        raise dash.exceptions.PreventUpdate
    value_to_remove = triggered["index"]
    updated = [v for v in (current_values or []) if str(v) != value_to_remove]
    field   = (context or {}).get("field", "")
    if not updated:
        return updated, True, html.P("No values selected.",
                                     className="text-muted fst-italic",
                                     style={"fontSize": "12px"})
    items = [
        html.Div([
            html.Span(str(v), style={"fontSize": "12px", "flex": "1"}),
            dbc.Button("×", id={"type": "remove-selected-value", "index": str(v)},
                      color="danger", size="sm", outline=True,
                      style={"fontSize": "11px", "padding": "1px 6px"}),
        ], className="d-flex align-items-center justify-content-between mb-1 px-2 py-1",
           style={"backgroundColor": "#f8f9fa", "borderRadius": "4px",
                  "border": "1px solid #dee2e6"})
        for v in updated
    ]
    return updated, True, html.Div([
        html.Div(f"{len(updated)} value{'s' if len(updated) != 1 else ''} selected for {field}",
                className="text-muted mb-2", style={"fontSize": "11px"}),
        html.Div(items)
    ])


# ─────────────────────────────────────────────
# Worksheet Settings
# ─────────────────────────────────────────────

@app.callback(
    Output("ws-settings-modal",     "is_open"),
    Output("ws-settings-header",    "children"),
    Output("ws-col-grand-total",    "value"),
    Output("ws-row-grand-total",    "value"),
    Input("ws-settings-payload",    "data"),
    Input("ws-settings-cancel-btn", "n_clicks"),
    Input("ws-settings-apply-btn",  "n_clicks"),
    State("ws-settings-store",      "data"),
    prevent_initial_call=True
)
def toggle_ws_settings(payload, cancel, apply_clicks, ws_settings):
    triggered = ctx.triggered_id
    if triggered in ("ws-settings-cancel-btn", "ws-settings-apply-btn"):
        return False, dash.no_update, dash.no_update, dash.no_update
    if triggered == "ws-settings-payload" and payload:
        ws_key   = payload.replace(" ", "_")
        settings = (ws_settings or {}).get(ws_key, {})
        return (True, f"⚙️ {payload} — Settings",
                settings.get("col_grand_total", "last"),
                settings.get("row_grand_total", "first"))
    raise dash.exceptions.PreventUpdate


@app.callback(
    Output("ws-settings-store", "data"),
    Output({"type": "data-table-container", "index": ALL}, "children",
           allow_duplicate=True),
    Input("ws-settings-apply-btn", "n_clicks"),
    State("ws-settings-payload",   "data"),
    State("ws-col-grand-total",    "value"),
    State("ws-row-grand-total",    "value"),
    State("ws-settings-store",     "data"),
    State("worksheet-store",       "data"),
    State("worksheet-tabs",        "value"),
    prevent_initial_call=True
)
def save_ws_settings(n_clicks, payload, col_total, row_total,
                     ws_settings, worksheets, active_tab):
    if not n_clicks or not payload:
        raise dash.exceptions.PreventUpdate
    ws_key   = payload.replace(" ", "_")
    settings = dict(ws_settings or {})
    settings[ws_key] = {"col_grand_total": col_total, "row_grand_total": row_total}
    try:
        cfg = load_config()
        cfg["ws_settings"] = settings
        save_config(cfg)
    except Exception as e:
        print(f"Could not save ws_settings: {e}")

    results = []
    for w in (worksheets or []):
        wk = w.replace(" ", "_")
        if wk != (active_tab or "").replace(" ", "_"):
            results.append(dash.no_update)
            continue
        ws_cfg       = settings.get(wk, {})
        results_dir  = get_base_dir() / "data" / "results"
        parquet_path = results_dir / f"{wk}.parquet"
        meta_path    = results_dir / f"{wk}.meta.json"
        if not parquet_path.exists() or not meta_path.exists():
            results.append(dash.no_update)
            continue
        try:
            df = pd.read_parquet(parquet_path)
            with open(meta_path) as mf:
                meta = json.load(mf)
            meta_rows = meta.get("rows", [])
            meta_cols = meta.get("cols", [])
            if meta_cols and meta_rows:
                df = apply_pivot(df, meta_rows, meta_cols,
                                col_total=ws_cfg.get("col_grand_total", "last"),
                                row_total=ws_cfg.get("row_grand_total", "first"))
            display_df = apply_row_blanking(df, meta_rows)
            table      = build_html_table(df, display_df=display_df, rows=meta_rows)
            results.append(html.Div([
                html.Div([
                    html.Span(f"{len(display_df):,} rows returned",
                             className="text-muted", style={"fontSize": "11px"}),
                    html.Div([
                        dbc.Button(
                            [html.I(className="bi bi-file-earmark-excel me-1"), "Spreadsheet"],
                            id={"type": "export-btn", "index": wk},
                            color="success", size="sm", outline=True,
                            style={"fontSize": "11px", "padding": "1px 8px"}
                        ),
                        dbc.Button(
                            [html.I(className="bi bi-file-earmark-excel me-1"), "Crosstab"],
                            id={"type": "export-crosstab-btn", "index": wk},
                            color="primary", size="sm", outline=True,
                            style={"fontSize": "11px", "padding": "1px 8px"}
                        ),
                    ], className="d-flex gap-1"),
                ], className="d-flex align-items-center justify-content-between px-2 py-1",
                   style={"borderBottom": "1px solid #eee"}),
                table
            ]))
        except Exception as e:
            print(f"Could not update grand total: {e}")
            results.append(dash.no_update)
    while len(results) < len(worksheets or []):
        results.append(dash.no_update)
    return settings, results[:len(worksheets or [])]


@app.callback(
    Output("query-running-modal", "is_open", allow_duplicate=True),
    Input({"type": "data-table-container", "index": ALL}, "children"),
    prevent_initial_call=True
)
def hide_query_modal(children):
    return False


# ─────────────────────────────────────────────
# Render worksheets
# ─────────────────────────────────────────────

@app.callback(
    Output("worksheet-content", "children", allow_duplicate=True),
    Input("worksheet-store", "data"),
    State("worksheet-tabs", "value"),
    prevent_initial_call='initial_duplicate'
)
def render_all_worksheets(worksheets, active_tab):
    if not worksheets:
        worksheets = ["Worksheet 1"]
    if not active_tab or active_tab not in worksheets:
        active_tab = worksheets[0]
    return [build_ws_wrapper(w, display="block" if w == active_tab else "none")
            for w in worksheets]


@app.callback(
    Output("worksheet-tabs", "children", allow_duplicate=True),
    Output("worksheet-tabs", "value",    allow_duplicate=True),
    Input("worksheet-store", "data"),
    State("worksheet-tabs", "value"),
    prevent_initial_call='initial_duplicate'
)
def restore_tabs(worksheets, current_value):
    if not worksheets:
        worksheets = ["Worksheet 1"]
    tabs   = [dcc.Tab(label=w, value=w) for w in worksheets]
    active = current_value if current_value in worksheets else worksheets[0]
    return tabs, active


@app.callback(
    Output({"type": "ws-rows",          "index": ALL}, "data", allow_duplicate=True),
    Output({"type": "ws-cols",          "index": ALL}, "data", allow_duplicate=True),
    Output({"type": "ws-filters",       "index": ALL}, "data", allow_duplicate=True),
    Output({"type": "ws-field-filters", "index": ALL}, "data", allow_duplicate=True),
    Output({"type": "ws-date-formats",  "index": ALL}, "data", allow_duplicate=True),
    Output({"type": "ws-measure",       "index": ALL}, "data", allow_duplicate=True),
    Input("worksheet-store", "data"),
    prevent_initial_call='initial_duplicate'
)
def restore_ws_state(worksheets):
    if not worksheets:
        worksheets = ["Worksheet 1"]
    cfg      = load_config()
    ws_state = cfg.get("ws_state", {})
    rows_out, cols_out, filters_out = [], [], []
    field_filters_out, date_formats_out, measure_out = [], [], []
    for w in worksheets:
        ws_key        = w.replace(" ", "_")
        state         = ws_state.get(ws_key, {})
        field_filters = state.get("field_filters", {})
        # Strip \r from any saved filter values
        field_filters = {
            k: ({"values": [v.strip() for v in vals.get("values", [])],
                 "exclude": vals.get("exclude", False)}
                if isinstance(vals, dict)
                else [v.strip() for v in vals])
            for k, vals in field_filters.items()
        }
        saved_filters = state.get("filters", [])
        if not saved_filters and field_filters:
            saved_filters = list(set(k.split("|")[-1] for k in field_filters.keys()))
        rows_out.append(state.get("rows", []))
        cols_out.append(state.get("cols", []))
        filters_out.append(saved_filters)
        field_filters_out.append(field_filters)
        date_formats_out.append(state.get("date_formats", {}))
        measure_out.append(state.get("measure", "COUNTD_USAGE_ID"))
    return (rows_out, cols_out, filters_out,
            field_filters_out, date_formats_out, measure_out)


@app.callback(
    Output("global-calculations", "data", allow_duplicate=True),
    Output({"type": "left-panel", "index": ALL}, "children", allow_duplicate=True),
    Input("worksheet-store", "data"),
    State({"type": "left-panel", "index": ALL}, "id"),
    prevent_initial_call='initial_duplicate'
)
def restore_calculations(worksheets, panel_ids):
    calcs = load_config().get("global_calculations", {})
    if not calcs:
        raise dash.exceptions.PreventUpdate
    return calcs, [build_table_panel(pid["index"], calcs) for pid in panel_ids]


@app.callback(
    Output("ws-settings-store", "data", allow_duplicate=True),
    Input("worksheet-store", "data"),
    prevent_initial_call='initial_duplicate'
)
def restore_ws_settings(worksheets):
    return load_config().get("ws_settings", {})


@app.callback(
    Output({"type": "ws-wrapper", "index": ALL}, "style"),
    Output({"type": "data-table-container", "index": ALL}, "children",
           allow_duplicate=True),
    Input("worksheet-tabs", "value"),
    State("worksheet-store", "data"),
    State({"type": "data-table-container", "index": ALL}, "children"),
    prevent_initial_call=True
)
def show_active_worksheet(active_tab, worksheets, current_children):
    if not worksheets:
        worksheets = ["Worksheet 1"]
    styles = [{"height": "100%", "display": "block" if w == active_tab else "none"}
              for w in worksheets]

    # Check if the active tab needs lazy loading
    table_results = [dash.no_update] * len(worksheets)
    active_key = active_tab.replace(" ", "_")

    for i, w in enumerate(worksheets):
        ws_key = w.replace(" ", "_")
        if ws_key != active_key:
            continue
        # Check if current content is a lazy-load placeholder
        child = current_children[i] if i < len(current_children) else None
        if child is None:
            continue
        # Detect placeholder by checking if it's a dict with lazy-load-pending type
        is_pending = False
        try:
            if isinstance(child, dict):
                props = child.get("props", {})
                child_id = props.get("id", {})
                if isinstance(child_id, dict) and child_id.get("type") == "lazy-load-pending":
                    is_pending = True
        except Exception:
            pass

        if is_pending:
            # Build the table now
            table_results[i] = _build_saved_result(ws_key)

    return styles, table_results

def _build_saved_result(ws_key):
    """Build the HTML table for a single saved worksheet result."""
    cfg = load_config()
    ws_settings  = cfg.get("ws_settings", {})
    results_dir  = get_base_dir() / "data" / "results"
    parquet_path = results_dir / f"{ws_key}.parquet"
    meta_path    = results_dir / f"{ws_key}.meta.json"

    if not parquet_path.exists() or not meta_path.exists():
        return dash.no_update

    try:
        df = pd.read_parquet(parquet_path)
        with open(meta_path) as f:
            meta = json.load(f)
        rows = meta.get("rows", [])
        cols = meta.get("cols", [])

        if cols and rows and "Count" in df.columns:
            ws_cfg = ws_settings.get(ws_key, {})

            df_summary = None
            summary_path = results_dir / f"{ws_key}.summary.parquet"
            if summary_path.exists():
                try:
                    df_summary = pd.read_parquet(summary_path)
                except Exception:
                    pass

            df = apply_pivot(df, rows, cols,
                             col_total=ws_cfg.get("col_grand_total", "last"),
                             row_total=ws_cfg.get("row_grand_total", "first"),
                             df_summary=df_summary)

        display_df = apply_row_blanking(df, rows)
        table      = build_html_table(df, display_df=display_df, rows=rows)

        return html.Div([
            html.Div([
                html.Span(f"{len(display_df):,} rows — saved result",
                         className="text-muted fst-italic",
                         style={"fontSize": "11px"}),
                html.Div([
                    dbc.Button(
                        [html.I(className="bi bi-file-earmark-excel me-1"), "Spreadsheet"],
                        id={"type": "export-btn", "index": ws_key},
                        color="success", size="sm", outline=True,
                        style={"fontSize": "11px", "padding": "1px 8px"}
                    ),
                    dbc.Button(
                        [html.I(className="bi bi-file-earmark-excel me-1"), "Crosstab"],
                        id={"type": "export-crosstab-btn", "index": ws_key},
                        color="primary", size="sm", outline=True,
                        style={"fontSize": "11px", "padding": "1px 8px"}
                    ),
                ], className="d-flex gap-1"),
            ], className="d-flex align-items-center justify-content-between px-2 py-1",
               style={"borderBottom": "1px solid #eee"}),
            table
        ], style={"display": "flex", "flexDirection": "column",
                  "height": "100%", "minHeight": "0", "overflow": "hidden"})
    except Exception as e:
        print(f"Could not build saved result for {ws_key}: {e}")
        return dash.no_update

# ─────────────────────────────────────────────
# Worksheet tab management
# ─────────────────────────────────────────────

@app.callback(
    Output("new-worksheet-modal", "is_open"),
    Output("new-worksheet-name",  "value", allow_duplicate=True),
    Input("add-worksheet-btn",    "n_clicks"),
    Input("cancel-worksheet-btn", "n_clicks"),
    State("new-worksheet-modal",  "is_open"),
    prevent_initial_call=True
)
def toggle_modal(add, cancel, is_open):
    triggered = ctx.triggered_id
    if triggered == "add-worksheet-btn":
        return True, ""
    if triggered == "cancel-worksheet-btn":
        return False, ""
    raise dash.exceptions.PreventUpdate


@app.callback(
    Output("worksheet-store",      "data",    allow_duplicate=True),
    Output("new-worksheet-name",   "value",   allow_duplicate=True),
    Output("new-worksheet-modal",  "is_open", allow_duplicate=True),
    Output("worksheet-tabs",       "value",   allow_duplicate=True),
    Input("confirm-worksheet-btn",  "n_clicks"),
    Input("rename-worksheet-input", "data"),
    State("new-worksheet-name",     "value"),
    State("worksheet-store",        "data"),
    State("worksheet-tabs",         "value"),
    prevent_initial_call=True
)
def manage_worksheets(confirm_clicks, rename_value, new_name, worksheets, active_tab):
    triggered = ctx.triggered_id
    if triggered == "rename-worksheet-input" and rename_value:
        try:
            data       = json.loads(rename_value)
            old_name   = data.get("from", "")
            new_name_r = data.get("to", "").strip()
            if not new_name_r or old_name not in worksheets:
                raise dash.exceptions.PreventUpdate
            if new_name_r in worksheets:
                new_name_r = f"{new_name_r} (2)"
            updated = [new_name_r if w == old_name else w for w in worksheets]
            _rename_ws_state(old_name, new_name_r)
            _save_worksheets(updated)
            return updated, dash.no_update, dash.no_update, dash.no_update
        except dash.exceptions.PreventUpdate:
            raise
        except Exception as e:
            print(f"Rename error: {e}")
            raise dash.exceptions.PreventUpdate
    if triggered == "confirm-worksheet-btn":
        new_name = (new_name or "").strip() or f"Worksheet {len(worksheets) + 1}"
        if new_name in worksheets:
            new_name = f"{new_name} (2)"
        updated = worksheets + [new_name]
        _save_worksheets(updated)
        return updated, "", False, new_name
    raise dash.exceptions.PreventUpdate


@app.callback(
    Output("delete-worksheet-btn", "style"),
    Input("worksheet-tabs",  "value"),
    State("worksheet-store", "data"),
)
def show_delete_btn(active_tab, worksheets):
    return {"display": "none"}


# ─────────────────────────────────────────────
# Duplicate worksheet
# ─────────────────────────────────────────────

@app.callback(
    Output("worksheet-store", "data",  allow_duplicate=True),
    Output("worksheet-tabs",  "value", allow_duplicate=True),
    Input("dupe-payload", "data"),
    State("worksheet-store", "data"),
    prevent_initial_call=True
)
def duplicate_worksheet(payload, worksheets):
    if not payload:
        raise dash.exceptions.PreventUpdate
    src_name = payload.strip()
    if src_name not in (worksheets or []):
        raise dash.exceptions.PreventUpdate
    base     = f"{src_name} (2)"
    new_name = base
    counter  = 2
    while new_name in worksheets:
        counter += 1
        new_name = f"{src_name} ({counter})"
    updated = list(worksheets) + [new_name]
    _duplicate_ws_state(src_name, new_name)
    _save_worksheets(updated)
    return updated, new_name


# ─────────────────────────────────────────────
# Delete worksheet
# ─────────────────────────────────────────────

@app.callback(
    Output("delete-ws-confirm-modal", "is_open"),
    Output("delete-ws-confirm-body",  "children"),
    Output("delete-ws-pending",       "data"),
    Input("delete-ws-payload",        "data"),
    prevent_initial_call=True
)
def confirm_delete_worksheet(payload):
    if not payload:
        raise dash.exceptions.PreventUpdate
    name = payload.strip()
    return (
        True,
        f"Are you sure you want to delete \"{name}\"? This cannot be undone.",
        name,
    )


@app.callback(
    Output("worksheet-store", "data",             allow_duplicate=True),
    Output("worksheet-tabs",  "value",            allow_duplicate=True),
    Output("delete-ws-confirm-modal", "is_open",  allow_duplicate=True),
    Input("delete-ws-confirm-btn",  "n_clicks"),
    Input("delete-ws-cancel-btn",   "n_clicks"),
    State("delete-ws-pending",      "data"),
    State("worksheet-store",        "data"),
    State("worksheet-tabs",         "value"),
    prevent_initial_call=True
)
def execute_delete_worksheet(confirm, cancel, name, worksheets, active_tab):
    triggered = ctx.triggered_id
    if triggered == "delete-ws-cancel-btn":
        return dash.no_update, dash.no_update, False
    if triggered != "delete-ws-confirm-btn" or not name:
        raise dash.exceptions.PreventUpdate
    name = name.strip()
    if name not in (worksheets or []):
        raise dash.exceptions.PreventUpdate

    # Clean up files
    try:
        ws_key = name.replace(" ", "_")
        results_dir = get_base_dir() / "data" / "results"
        for ext in [".parquet", ".meta.json", ".summary.parquet"]:
            p = results_dir / f"{ws_key}{ext}"
            if p.exists():
                p.unlink()
        cfg = load_config()
        for key in ["ws_state", "ws_settings"]:
            store = cfg.get(key, {})
            store.pop(ws_key, None)
            cfg[key] = store
        save_config(cfg)
    except Exception as e:
        print(f"Could not clean up deleted worksheet: {e}")

    # If it's the last worksheet, replace with a blank one
    if len(worksheets) <= 1:
        updated = ["Worksheet 1"]
        _save_worksheets(updated)
        return updated, "Worksheet 1", False

    updated = [w for w in worksheets if w != name]
    _save_worksheets(updated)
    new_active = active_tab
    if active_tab == name:
        idx = worksheets.index(name)
        new_active = updated[min(idx, len(updated) - 1)]
    return updated, new_active, False


# ─────────────────────────────────────────────
# Run Query button state — clientside
# ─────────────────────────────────────────────

app.clientside_callback(
    """
function(rows_list, cols_list, field_filters_list, measure_list, last_run, active_tab) {
    var n = rows_list.length;
    var disabled = [], colors = [], children = [];
    var active_key = (active_tab || "").replace(/ /g, "_");

    for (var i = 0; i < n; i++) {
        var rows = rows_list[i] || [];
        var cols = cols_list[i] || [];
        disabled.push(!(rows.length > 0 || cols.length > 0));

        var inputs = dash_clientside.callback_context.inputs_list;
        var ws_key = inputs && inputs[0] && inputs[0][i] ?
                     inputs[0][i].id.index : "";

        if (ws_key !== active_key) {
            colors.push(window.dash_clientside.no_update);
            children.push(window.dash_clientside.no_update);
            continue;
        }

        var last = (last_run && last_run[ws_key]) ? last_run[ws_key] : null;
        if (!last) {
            colors.push("primary");
            children.push("Run Query");
            continue;
        }

        var ff = field_filters_list[i] || {};
        var measure = measure_list[i] || "COUNTD_USAGE_ID";

        var is_stale = (
            JSON.stringify(rows)    !== JSON.stringify(last.rows    || []) ||
            JSON.stringify(cols)    !== JSON.stringify(last.cols    || []) ||
            JSON.stringify(ff)      !== JSON.stringify(last.field_filters || {}) ||
            measure                 !== (last.measure || "COUNTD_USAGE_ID")
        );

        if (is_stale) {
            colors.push("warning");
            children.push("Run Query ●");
        } else {
            colors.push("primary");
            children.push("Run Query");
        }
    }
    return [disabled, colors, children];
}
""",
    Output({"type": "run-query-btn", "index": ALL}, "disabled"),
    Output({"type": "run-query-btn", "index": ALL}, "color"),
    Output({"type": "run-query-btn", "index": ALL}, "children"),
    Input({"type": "ws-rows",         "index": ALL}, "data"),
    Input({"type": "ws-cols",         "index": ALL}, "data"),
    Input({"type": "ws-field-filters","index": ALL}, "data"),
    Input({"type": "ws-measure",      "index": ALL}, "data"),
    Input("ws-last-run-state", "data"),
    State("worksheet-tabs", "value"),
)


# ─────────────────────────────────────────────
# Shelves — R/C/F buttons
# ─────────────────────────────────────────────

@app.callback(
    Output({"type": "ws-rows", "index": ALL}, "data"),
    Input({"type": "add-rows-btn", "index": ALL}, "n_clicks"),
    State({"type": "ws-rows",      "index": ALL}, "data"),
    prevent_initial_call=True
)
def add_to_rows(n_clicks, current_rows):
    if not any(n for n in n_clicks if n):
        raise dash.exceptions.PreventUpdate
    triggered = ctx.triggered_id
    if triggered is None:
        raise dash.exceptions.PreventUpdate
    field    = triggered["index"].split("|")[-1]
    ws_key   = triggered["index"].split("|")[0]
    ws_index = next((i for i, item in enumerate(ctx.states_list[0])
                     if item["id"]["index"] == ws_key), 0)
    updated  = list(current_rows[ws_index] or [])
    if field not in updated:
        updated.append(field)
    return [updated if i == ws_index else dash.no_update
            for i in range(len(current_rows))]


@app.callback(
    Output({"type": "ws-rows", "index": ALL}, "data", allow_duplicate=True),
    Input({"type": "remove-rows-btn", "index": ALL}, "n_clicks"),
    State({"type": "ws-rows",         "index": ALL}, "data"),
    prevent_initial_call=True
)
def remove_from_rows(n_clicks, current_rows):
    if not any(n for n in n_clicks if n):
        raise dash.exceptions.PreventUpdate
    triggered = ctx.triggered_id
    if triggered is None:
        raise dash.exceptions.PreventUpdate
    parts    = triggered["index"].split("|")
    ws_key   = parts[0]
    field    = "|".join(parts[1:])
    ws_index = next((i for i, item in enumerate(ctx.states_list[0])
                     if item["id"]["index"] == ws_key), 0)
    updated  = [f for f in (current_rows[ws_index] or []) if f != field]
    clear_saved_result(ws_key)
    return [updated if i == ws_index else dash.no_update
            for i in range(len(current_rows))]


@app.callback(
    Output({"type": "ws-cols", "index": ALL}, "data"),
    Input({"type": "add-cols-btn", "index": ALL}, "n_clicks"),
    State({"type": "ws-cols",      "index": ALL}, "data"),
    prevent_initial_call=True
)
def add_to_cols(n_clicks, current_cols):
    if not any(n for n in n_clicks if n):
        raise dash.exceptions.PreventUpdate
    triggered = ctx.triggered_id
    if triggered is None:
        raise dash.exceptions.PreventUpdate
    field    = triggered["index"].split("|")[-1]
    ws_key   = triggered["index"].split("|")[0]
    ws_index = next((i for i, item in enumerate(ctx.states_list[0])
                     if item["id"]["index"] == ws_key), 0)
    updated  = list(current_cols[ws_index] or [])
    if field not in updated:
        updated.append(field)
    return [updated if i == ws_index else dash.no_update
            for i in range(len(current_cols))]


@app.callback(
    Output({"type": "ws-cols", "index": ALL}, "data", allow_duplicate=True),
    Input({"type": "remove-cols-btn", "index": ALL}, "n_clicks"),
    State({"type": "ws-cols",         "index": ALL}, "data"),
    prevent_initial_call=True
)
def remove_from_cols(n_clicks, current_cols):
    if not any(n for n in n_clicks if n):
        raise dash.exceptions.PreventUpdate
    triggered = ctx.triggered_id
    if triggered is None:
        raise dash.exceptions.PreventUpdate
    parts    = triggered["index"].split("|")
    ws_key   = parts[0]
    field    = "|".join(parts[1:])
    ws_index = next((i for i, item in enumerate(ctx.states_list[0])
                     if item["id"]["index"] == ws_key), 0)
    updated  = [f for f in (current_cols[ws_index] or []) if f != field]
    clear_saved_result(ws_key)
    return [updated if i == ws_index else dash.no_update
            for i in range(len(current_cols))]


# ─────────────────────────────────────────────
# Shelf renders
# ─────────────────────────────────────────────

@app.callback(
    Output({"type": "rows-shelf", "index": ALL}, "children"),
    Input({"type": "ws-rows",          "index": ALL}, "data"),
    Input({"type": "ws-field-filters", "index": ALL}, "data"),
    Input({"type": "ws-date-formats",  "index": ALL}, "data"),
)
def render_rows_shelf(rows_data, field_filters_data, date_formats_data):
    results = []
    for idx in range(len(rows_data)):
        worksheet_id  = ctx.inputs_list[0][idx]["id"]["index"]
        rows          = rows_data[idx] or []
        field_filters = field_filters_data[idx] if idx < len(field_filters_data) else {}
        date_formats  = date_formats_data[idx]  if idx < len(date_formats_data)  else {}
        results.append(_render_rows_items(rows, field_filters, date_formats, worksheet_id))
    return results


@app.callback(
    Output({"type": "cols-shelf", "index": ALL}, "children"),
    Input({"type": "ws-cols",          "index": ALL}, "data"),
    Input({"type": "ws-field-filters", "index": ALL}, "data"),
    Input({"type": "ws-date-formats",  "index": ALL}, "data"),
)
def render_cols_shelf(cols_data, field_filters_data, date_formats_data):
    results = []
    for idx in range(len(cols_data)):
        worksheet_id  = ctx.inputs_list[0][idx]["id"]["index"]
        cols          = cols_data[idx] or []
        field_filters = field_filters_data[idx] if idx < len(field_filters_data) else {}
        date_formats  = date_formats_data[idx]  if idx < len(date_formats_data)  else {}
        results.append(_render_cols_items(cols, field_filters, date_formats, worksheet_id))
    return results


# ─────────────────────────────────────────────
# Drag and drop
# ─────────────────────────────────────────────

@app.callback(
    Output({"type": "ws-rows",    "index": ALL}, "data", allow_duplicate=True),
    Output({"type": "ws-cols",    "index": ALL}, "data", allow_duplicate=True),
    Output({"type": "ws-filters", "index": ALL}, "data", allow_duplicate=True),
    Output({"type": "rows-shelf", "index": ALL}, "children", allow_duplicate=True),
    Output({"type": "cols-shelf", "index": ALL}, "children", allow_duplicate=True),
    Output("badge-context-payload", "data", allow_duplicate=True),
    Input("drop-payload", "data"),
    State({"type": "ws-rows",          "index": ALL}, "data"),
    State({"type": "ws-cols",          "index": ALL}, "data"),
    State({"type": "ws-filters",       "index": ALL}, "data"),
    State({"type": "ws-field-filters", "index": ALL}, "data"),
    State({"type": "ws-date-formats",  "index": ALL}, "data"),
    prevent_initial_call=True
)
def handle_drop_and_render(payload, rows_data, cols_data, filters_data,
                            field_filters_data, date_formats_data):
    if not payload:
        raise dash.exceptions.PreventUpdate
    try:
        data = json.loads(payload)
        data.pop("_t", None)
        field         = data.get("field", "")
        shelf         = data.get("shelf", "")
        worksheet     = data.get("worksheet", "")
        source        = data.get("source", "panel")
        insert_before = data.get("insert_before")
        insert_after  = data.get("insert_after")
    except Exception:
        raise dash.exceptions.PreventUpdate

    if not field or not shelf or not worksheet:
        raise dash.exceptions.PreventUpdate

    n        = len(rows_data)
    ws_index = next((i for i, item in enumerate(ctx.states_list[0])
                     if item["id"]["index"] == worksheet), None)
    if ws_index is None:
        raise dash.exceptions.PreventUpdate

    rows    = list(rows_data[ws_index]    or [])
    cols    = list(cols_data[ws_index]    or [])
    filters = list(filters_data[ws_index] or [])

    if source == "rows"    and field in rows:    rows.remove(field)
    if source == "cols"    and field in cols:    cols.remove(field)
    if source == "filters" and field in filters: filters.remove(field)

    def insert_into(lst, f, before=None, after=None):
        if f in lst: lst.remove(f)
        if before and before in lst:
            lst.insert(lst.index(before), f)
        elif after and after in lst:
            lst.insert(lst.index(after) + 1, f)
        else:
            lst.append(f)
        return lst

    if shelf == "rows":
        rows    = insert_into(rows,    field, insert_before, insert_after)
    elif shelf == "cols":
        cols    = insert_into(cols,    field, insert_before, insert_after)
    elif shelf == "filters":
        filters = insert_into(filters, field, insert_before, insert_after)

    field_filters = field_filters_data[ws_index] if field_filters_data else {}
    date_formats  = date_formats_data[ws_index]  if date_formats_data  else {}
    worksheet_id  = ctx.states_list[0][ws_index]["id"]["index"]

    # If dropped on filters shelf, auto-open the filter dialog
    filter_payload = dash.no_update
    if shelf == "filters":
        filter_payload = json.dumps({
            "field": field,
            "worksheet": worksheet,
            "shelf": "filters",
            "tab": "filter"
        })

    return (
        [rows    if i == ws_index else dash.no_update for i in range(n)],
        [cols    if i == ws_index else dash.no_update for i in range(n)],
        [filters if i == ws_index else dash.no_update for i in range(n)],
        [_render_rows_items(rows, field_filters, date_formats, worksheet_id)
         if i == ws_index else dash.no_update for i in range(n)],
        [_render_cols_items(cols, field_filters, date_formats, worksheet_id)
         if i == ws_index else dash.no_update for i in range(n)],
        filter_payload,
    )


# ─────────────────────────────────────────────
# Filter panel
# ─────────────────────────────────────────────

@app.callback(
    Output({"type": "filter-panel", "index": ALL}, "children"),
    Input({"type": "ws-filters",      "index": ALL}, "data"),
    Input({"type": "ws-field-filters","index": ALL}, "data"),
    Input({"type": "ws-date-formats", "index": ALL}, "data"),
)
def render_filter_panel(filters_data, field_filters_data, date_formats_data):
    results = []
    for idx, filters_item in enumerate(filters_data):
        filters       = filters_item or []
        field_filters = field_filters_data[idx] if idx < len(field_filters_data) else {}
        date_formats  = (date_formats_data[idx]
                         if date_formats_data and idx < len(date_formats_data) else {})
        worksheet_id  = ctx.inputs_list[0][idx]["id"]["index"]
        if not filters:
            results.append(html.Span("Drag fields here or use F button.",
                           className="text-muted", style={"fontSize": "11px"}))
            continue
        items = []
        for f in filters:
            saved_ff = field_filters.get(f"filters|{f}", field_filters.get(f, []))
            if isinstance(saved_ff, dict):
                values          = saved_ff.get("values", [])
                exclude_active  = saved_ff.get("exclude", False)
            else:
                values         = saved_ff
                exclude_active = False
            active = bool(values)
            fmt    = _get_fmt(date_formats, "filters", f)
            if values:
                prefix = "NOT: " if exclude_active else ""
                desc   = prefix + ", ".join(str(v) for v in values[:3])
                if len(values) > 3:
                    desc += f" +{len(values)-3}"
            else:
                desc = None
            fmt_label = f if fmt == "none" else f"{f} [{FMT_SHORT.get(fmt, fmt)}]"
            items.append(html.Div([
                html.Div([
                    dbc.Button(
                        [html.Span(fmt_label, style={"fontSize": "11px"}),
                         html.I(className="bi bi-funnel-fill ms-1",
                               style={"fontSize": "9px",
                                      "color": "yellow" if active else "white"})],
                        id={"type": "field-filter-open",
                            "index": f"{worksheet_id}|filters|{f}"},
                        color="secondary", size="sm", className="text-start",
                        style={"fontSize": "11px", "padding": "3px 8px",
                               "borderRadius": "4px 0 0 4px", "borderRight": "none"}
                    ),
                    dbc.Button("×",
                        id={"type": "remove-filter-btn",
                            "index": f"{worksheet_id}|{f}"},            color="secondary", size="sm",
                        style={"fontSize": "13px", "padding": "2px 5px",
                               "borderRadius": "0 4px 4px 0"}),
                ], className="d-flex mb-0 draggable-badge", draggable="true",
                   **{"data-field": f, "data-source-shelf": "filters",
                      "data-worksheet": worksheet_id, "data-fmt": fmt}),
                html.Div(desc,
                    style={"fontSize": "10px", "color": "#984b0c",
                           "paddingLeft": "4px", "marginBottom": "4px",
                           "whiteSpace": "nowrap", "overflow": "hidden",
                           "textOverflow": "ellipsis", "maxWidth": "190px",
                           "fontStyle": "italic"}
                ) if desc else html.Div(style={"marginBottom": "4px"}),
            ]))
        results.append(items)
    return results


@app.callback(
    Output({"type": "ws-filters",       "index": ALL}, "data", allow_duplicate=True),
    Output({"type": "ws-field-filters", "index": ALL}, "data", allow_duplicate=True),
    Input({"type": "remove-filter-btn", "index": ALL}, "n_clicks"),
    State({"type": "ws-filters",        "index": ALL}, "data"),
    State({"type": "ws-field-filters",  "index": ALL}, "data"),
    prevent_initial_call=True
)
def remove_from_filters(n_clicks, current_filters, current_field_filters):
    if not any(n for n in n_clicks if n):
        raise dash.exceptions.PreventUpdate
    triggered = ctx.triggered_id
    if triggered is None:
        raise dash.exceptions.PreventUpdate
    parts    = triggered["index"].split("|")
    ws_key   = parts[0]
    field    = "|".join(parts[1:])
    ws_index = next((i for i, item in enumerate(ctx.states_list[0])
                     if item["id"]["index"] == ws_key), 0)
    updated_filters = [f for f in (current_filters[ws_index] or []) if f != field]
    # Also remove from field_filters
    updated_ff = dict(current_field_filters[ws_index] or {})
    updated_ff.pop(f"filters|{field}", None)
    updated_ff.pop(field, None)
    n = len(current_filters)
    return (
        [updated_filters if i == ws_index else dash.no_update for i in range(n)],
        [updated_ff      if i == ws_index else dash.no_update for i in range(n)],
    )

# ─────────────────────────────────────────────
# Values shelf
# ─────────────────────────────────────────────

@app.callback(
    Output({"type": "ws-measure", "index": ALL}, "data"),
    Input({"type": "measure-select", "index": ALL}, "value"),
    prevent_initial_call=True
)
def save_measure(values):
    triggered = ctx.triggered_id
    if triggered is None:
        raise dash.exceptions.PreventUpdate
    ti  = next((i for i, item in enumerate(ctx.inputs_list[0])
                if item["id"] == triggered), 0)
    val = values[ti] if values else "COUNTD_USAGE_ID"
    return [val if i == ti else dash.no_update for i in range(len(values))]


# ─────────────────────────────────────────────
# Badge right-click
# ─────────────────────────────────────────────

@app.callback(
    Output({"type": "ws-date-formats", "index": ALL}, "data", allow_duplicate=True),
    Output("perm-field-filter-modal",     "is_open",    allow_duplicate=True),
    Output("perm-field-filter-header",    "children",   allow_duplicate=True),
    Output("perm-field-filter-checklist", "value",      allow_duplicate=True),
    Output("perm-field-filter-summary",   "children",   allow_duplicate=True),
    Output("perm-field-context",          "data",       allow_duplicate=True),
    Output({"type": "ws-filters", "index": ALL}, "data", allow_duplicate=True),
    Output("perm-filter-exclude",         "value",      allow_duplicate=True),
    Output("perm-filter-sort",            "data",       allow_duplicate=True),
    Output("perm-filter-sort-btn",        "children",   allow_duplicate=True),
    Output("perm-filter-date-fmt",        "options",    allow_duplicate=True),
    Output("perm-filter-date-fmt",        "value",      allow_duplicate=True),
    Output("perm-filter-date-fmt",        "disabled",   allow_duplicate=True),
    Output("perm-filter-cascade",         "value",      allow_duplicate=True),
    Input("badge-context-payload", "data"),
    State({"type": "ws-field-filters", "index": ALL}, "data"),
    State({"type": "ws-date-formats",  "index": ALL}, "data"),
    State({"type": "ws-filters",       "index": ALL}, "data"),
    State("perm-filter-sort", "data"),
    prevent_initial_call=True
)
def handle_badge_context(payload, field_filters_data, date_formats_data, ws_filters_data, sort_data):
    if not payload:
        raise dash.exceptions.PreventUpdate
    try:
        data         = json.loads(payload)
        field        = data.get("field", "")
        worksheet_id = data.get("worksheet", "")
        shelf        = data.get("shelf", "rows")
        tab_request  = data.get("tab", "filter")
        new_fmt      = data.get("fmt", None)
    except Exception:
        raise dash.exceptions.PreventUpdate
    if not field or not worksheet_id:
        raise dash.exceptions.PreventUpdate
    ws_index = next((i for i, item in enumerate(ctx.states_list[0])
                     if item["id"]["index"] == worksheet_id), 0)
    field_filters = field_filters_data[ws_index] if field_filters_data else {}
    date_formats  = dict(date_formats_data[ws_index] if date_formats_data else {})
    num_fmt       = len(date_formats_data) if date_formats_data else 1
    num_filters   = len(ws_filters_data)   if ws_filters_data   else 1
    if tab_request == "format" and new_fmt is not None:
        date_formats[_fmt_key(shelf, field)] = new_fmt
        return (
            [date_formats if i == ws_index else dash.no_update for i in range(num_fmt)],
            dash.no_update, dash.no_update, dash.no_update,
            dash.no_update, dash.no_update,
            [dash.no_update] * num_filters,
            dash.no_update,
            dash.no_update, dash.no_update,
            dash.no_update, dash.no_update, dash.no_update,
            dash.no_update,
        )
    if tab_request == "remove":
        raise dash.exceptions.PreventUpdate
    # Always show filter shelf format when opening filter dialog
    current_fmt = _get_fmt(date_formats, "filters", field)
    saved_ff    = field_filters.get(f"filters|{field}", field_filters.get(field, []))
    if isinstance(saved_ff, dict):
        current_values  = saved_ff.get("values", [])
        current_exclude = saved_ff.get("exclude", False)
    else:
        current_values  = saved_ff if isinstance(saved_ff, list) else []
        current_exclude = False
    context = {"worksheet_id": worksheet_id, "field": field,
                "shelf": "filters", "fmt": current_fmt}
    summary = _build_filter_summary(current_values, current_exclude)

    is_date = field in DATE_FIELDS
    if is_date:
        fmt_options = [{"label": "None", "value": "none"}] + [
            {"label": opt["label"], "value": opt["value"]}
            for opt in DATE_FORMAT_OPTIONS
        ]
    else:
        fmt_options = [{"label": "None", "value": "none"}]

    sort_data = sort_data if isinstance(sort_data, dict) else {}
    field_sort = sort_data.get(field, "asc")
    field_sort_label = "Z→A" if field_sort == "desc" else "A→Z"

    return (
        [dash.no_update] * num_fmt,
        True,
        f"⚙ {field}",
        current_values,
        summary,
        context,
        [dash.no_update] * num_filters,
        current_exclude,
        sort_data, field_sort_label,
        fmt_options, current_fmt, not is_date,
        False,
    )


# ─────────────────────────────────────────────
# Filter modal — F button and shelf filter button
# ─────────────────────────────────────────────

@app.callback(
    Output("perm-field-filter-modal",     "is_open"),
    Output("perm-field-filter-header",    "children"),
    Output("perm-field-filter-checklist", "options"),
    Output("perm-field-filter-checklist", "value"),
    Output("perm-field-filter-summary",   "children"),
    Output("perm-field-context",          "data"),
    Output({"type": "ws-filters", "index": ALL}, "data"),
    Output("perm-filter-exclude",         "value"),
    Output("perm-filter-sort",            "data"),
    Output("perm-filter-sort-btn",        "children"),
    Output("perm-filter-date-fmt",        "options"),
    Output("perm-filter-date-fmt",        "value"),
    Output("perm-filter-date-fmt",        "disabled"),
    Output("perm-filter-cascade",         "value"),
    Input({"type": "add-filter-btn",   "index": ALL}, "n_clicks"),
    Input({"type": "field-filter-open","index": ALL}, "n_clicks"),
    State({"type": "ws-field-filters", "index": ALL}, "data"),
    State({"type": "ws-date-formats",  "index": ALL}, "data"),
    State({"type": "ws-filters",       "index": ALL}, "data"),
    State("perm-filter-sort",      "data"),
    prevent_initial_call=True
)
def open_field_filter(filter_clicks, shelf_clicks,
                      field_filters_data, date_formats_data, ws_filters_data, sort_data):
    triggered = ctx.triggered_id
    if triggered is None:
        raise dash.exceptions.PreventUpdate

    if isinstance(triggered, dict) and triggered.get("type") == "add-filter-btn":
        triggered_clicks = next(
            (n for i, n in enumerate(filter_clicks)
             if ctx.inputs_list[0][i]["id"] == triggered), None
        )
        if not triggered_clicks:
            raise dash.exceptions.PreventUpdate
        parts        = triggered["index"].split("|")
        worksheet_id = parts[0]
        if len(parts) >= 3 and parts[1] == "calc":
            field = "|".join(parts[2:])
        else:
            field = parts[-1]
        if field.startswith("calc_"):
            field = field[5:]
        shelf = "filters"

    elif isinstance(triggered, dict) and triggered.get("type") == "field-filter-open":
        triggered_clicks = next(
            (n for i, n in enumerate(shelf_clicks)
             if ctx.inputs_list[1][i]["id"] == triggered), None
        )
        if not triggered_clicks:
            raise dash.exceptions.PreventUpdate
        parts = triggered["index"].split("|")
        if len(parts) != 3:
            raise dash.exceptions.PreventUpdate
        worksheet_id, shelf, field = parts[0], parts[1], parts[2]
        if field.startswith("calc_"):
            field = field[5:]
    else:
        raise dash.exceptions.PreventUpdate

    ws_index       = next((i for i, item in enumerate(ctx.states_list[0])
                           if item["id"]["index"] == worksheet_id), 0)
    field_filters  = field_filters_data[ws_index] if field_filters_data else {}
    date_formats   = date_formats_data[ws_index]  if date_formats_data  else {}
    # Filter dialog always uses the filter shelf's format
    current_fmt    = _get_fmt(date_formats, "filters", field)
    filter_key     = f"filters|{field}"
    saved          = field_filters.get(filter_key, field_filters.get(field, {}))
    # Support both old format (list) and new format (dict with values+exclude)
    if isinstance(saved, list):
        current_values  = saved
        current_exclude = False
    elif isinstance(saved, dict):
        current_values  = saved.get("values", [])
        current_exclude = saved.get("exclude", False)
    else:
        current_values  = []
        current_exclude = False

    context        = {"worksheet_id": worksheet_id, "field": field,
                      "shelf": shelf, "fmt": current_fmt}
    num_filters    = len(ws_filters_data) if ws_filters_data else 1
    ws_filters_out = [dash.no_update] * num_filters
    summary = _build_filter_summary(current_values, current_exclude)
    cfg     = load_config()
    calcs   = cfg.get("global_calculations", {})
    options = _get_options_for_field(field, current_fmt, calcs)
    is_date = field in DATE_FIELDS
    if is_date:
        fmt_options = [{"label": "None", "value": "none"}] + [
            {"label": opt["label"], "value": opt["value"]}
            for opt in DATE_FORMAT_OPTIONS
        ]
    else:
        fmt_options = [{"label": "None", "value": "none"}]

    sort_data = sort_data if isinstance(sort_data, dict) else {}
    field_sort = sort_data.get(field, "asc")
    field_sort_label = "Z→A" if field_sort == "desc" else "A→Z"

    return (True, f"⚙ {field}", dash.no_update, current_values,
            summary, context, ws_filters_out, current_exclude,
            sort_data, field_sort_label,
            fmt_options, current_fmt, not is_date,
            False)
            

@app.callback(
    Output("perm-field-filter-checklist", "options"),
    Output("filter-pending-search",        "data"),
    Input("perm-filter-search-btn",       "n_clicks"),
    Input("perm-filter-clear-search-btn", "n_clicks"),
    Input("perm-filter-sort", "data"),
    Input("perm-field-context",           "data"),
    Input("perm-filter-cascade",          "value"),
    State("perm-filter-search",           "value"),
    State("perm-filter-sort",             "data"),
    State({"type": "ws-field-filters", "index": ALL}, "data"),
    prevent_initial_call=True
)
def search_filter_options(search_clicks, clear_clicks, sort_data,
                          context_input, cascade, search, sort_order,
                          field_filters_data):
    if not context_input:
        raise dash.exceptions.PreventUpdate

    field    = context_input.get("field", "")
    fmt      = context_input.get("fmt", "none")
    cfg      = load_config()
    calcs    = cfg.get("global_calculations", {})
    sort_order_data = sort_order if isinstance(sort_order, dict) else {}
    field = context_input.get("field", "")
    reverse = (sort_order_data.get(field, "asc") == "desc")
    triggered = ctx.triggered_id

    def get_full_sorted():
        full = list(_get_options_for_field(field, fmt, calcs))
        real = [o for o in full if o.get("value") not in
                ("__hint__", "__loading__", "__error__", "__none__")]
        real.sort(key=lambda o: str(o.get("label", "")).lower(), reverse=reverse)
        return real if real else \
               [{"label": "No values found", "value": "__none__", "disabled": True}]

    if triggered == "perm-field-context":
        if field in HIGH_CARDINALITY_FIELDS:
            return [], ""
        return get_full_sorted(), ""

    if triggered == "perm-filter-clear-search-btn":
        if field in HIGH_CARDINALITY_FIELDS:
            return [], ""
        return get_full_sorted(), ""

    if triggered == "perm-filter-cascade":
        if field in HIGH_CARDINALITY_FIELDS:
            search_lower = (search or "").strip().lower()
            if len(search_lower) >= 3:
                search_lower = search_lower.replace("'", "''")
                return [], search_lower
            return [], ""
        return get_full_sorted(), ""

    if triggered == "perm-filter-sort":
        if field in HIGH_CARDINALITY_FIELDS:
            search_lower = (search or "").strip().lower()
            if len(search_lower) >= 3:
                # Re-run the search with new sort order
                search_lower = search_lower.replace("'", "''")
                return [], search_lower
            return [], ""
        return get_full_sorted(), "" ""

    if triggered == "perm-filter-search-btn":
        search_lower = (search or "").strip().lower()
        if not search_lower:
            if field in HIGH_CARDINALITY_FIELDS:
                return [], ""
            return get_full_sorted(), ""

        if field in HIGH_CARDINALITY_FIELDS:
            if len(search_lower) < 3:
                return [], ""
            # Clear the list immediately, store the term for execute callback
            search_lower = search_lower.replace("'", "''")
            return [], search_lower

        # Normal cardinality — do it inline, no race risk
        full     = list(_get_options_for_field(field, fmt, calcs))
        import fnmatch
        if '*' in search_lower or '?' in search_lower:
            filtered = [o for o in full
                        if fnmatch.fnmatch(str(o.get("label", "")).lower(), search_lower)
                        and o.get("value") not in ("__loading__", "__error__",
                                                   "__hint__", "__none__")]
        else:
            filtered = [o for o in full
                        if search_lower in str(o.get("label", "")).lower()
                        and o.get("value") not in ("__loading__", "__error__",
                                                   "__hint__", "__none__")]
        filtered.sort(key=lambda o: str(o.get("label", "")), reverse=reverse)
        return filtered if filtered else \
               [{"label": "No matches found", "value": "__none__", "disabled": True}], ""

    raise dash.exceptions.PreventUpdate

@app.callback(
    Output("perm-field-filter-checklist", "options", allow_duplicate=True),
    Output("perm-field-context",          "data",    allow_duplicate=True),
    Input("perm-filter-date-fmt", "value"),
    State("perm-field-context",   "data"),
    State("perm-filter-sort",     "data"),
    prevent_initial_call=True
)
def update_options_on_fmt_change(new_fmt, context, sort_order):
    if not context or not context.get("field"):
        raise dash.exceptions.PreventUpdate
    field   = context["field"]
    if field not in DATE_FIELDS:
        raise dash.exceptions.PreventUpdate
    # Don't re-fetch if the format hasn't actually changed
    if new_fmt == context.get("fmt", "none"):
        raise dash.exceptions.PreventUpdate

    # Update context with new format
    updated_context = dict(context)
    updated_context["fmt"] = new_fmt

    reverse = (sort_order == "desc")
    cfg     = load_config()
    calcs   = cfg.get("global_calculations", {})
    full    = list(_get_options_for_field(field, new_fmt, calcs))
    real    = [o for o in full if o.get("value") not in
              ("__hint__", "__loading__", "__error__", "__none__")]
    real.sort(key=lambda o: str(o.get("label", "")), reverse=reverse)
    options = real if real else \
              [{"label": "No values found", "value": "__none__", "disabled": True}]

    return options, updated_context

@app.callback(
    Output("perm-field-filter-checklist", "options", allow_duplicate=True),
    Input("filter-pending-search", "data"),
    State("perm-field-context",    "data"),
    State("perm-filter-sort",      "data"),
    State("perm-filter-cascade",   "value"),
    State({"type": "ws-field-filters", "index": ALL}, "data"),
    prevent_initial_call=True
)
def execute_hc_search(search_lower, context_input, sort_order, cascade, field_filters_data):
    if not search_lower or not context_input:
        raise dash.exceptions.PreventUpdate
    field = context_input.get("field", "")
    if field not in HIGH_CARDINALITY_FIELDS:
        raise dash.exceptions.PreventUpdate

    search_lower = search_lower.replace("'", "''")
    sort_order_data = sort_order if isinstance(sort_order, dict) else {}
    reverse = (sort_order_data.get(field, "asc") == "desc")
    sort_dir = "DESC" if reverse else "ASC"

    if field in _STUDY_FILTER_FIELDS:
        parquet_src = f"read_parquet('{str(STUDY_PATH)}')"
    elif field in _USER_GROUP_FILTER_FIELDS:
        parquet_src = f"read_parquet('{str(USER_GROUP_PATH)}')"
    else:
        parquet_src = f"read_parquet('{str(USAGE_PATH)}')"

    # Build cascade filters
    where_extra = ""
    print(f"CASCADE DEBUG: field={field}, where_extra={where_extra}")
    if cascade and context_input.get("worksheet_id"):
        ws_id = context_input["worksheet_id"]
        # field_filters_data is a list — we need to match by worksheet index
        # The index positions correspond to worksheet order
        ws_idx = 0
        try:
            cfg_ws = load_config().get("worksheets", [])
            for i, w in enumerate(cfg_ws):
                if w.replace(" ", "_") == ws_id:
                    ws_idx = i
                    break
        except Exception:
            pass
        ff = field_filters_data[ws_idx] if field_filters_data else {}
        same_clauses, cross_fields = _build_cascade_where(ff, field)

        if same_clauses:
            where_extra = " AND " + " AND ".join(same_clauses)

        # Cross-table: e.g. filtering LONG_NAME (study) by CLIENT_NAME (usage)
        for cross in cross_fields:
            if cross["current_table"] == "study" and cross["filter_table"] == "usage":
                where_extra += f"""
                    AND STUDY_ID IN (
                        SELECT DISTINCT b.STUDY_ID
                        FROM read_parquet('{str(BRIDGE_PATH)}') b
                        JOIN read_parquet('{str(USAGE_PATH)}') u ON b.USAGE_ID = u.USAGE_ID
                        WHERE {cross['where_str']}
                    )"""
            elif cross["current_table"] == "usage" and cross["filter_table"] == "study":
                where_extra += f"""
                    AND USAGE_ID IN (
                        SELECT DISTINCT b.USAGE_ID
                        FROM read_parquet('{str(BRIDGE_PATH)}') b
                        JOIN read_parquet('{str(STUDY_PATH)}') s ON b.STUDY_ID = s.STUDY_ID
                        WHERE {cross['where_str']}
                    )"""
            elif cross["current_table"] == "study" and cross["filter_table"] == "user_group":
                where_extra += f"""
                    AND STUDY_ID IN (
                        SELECT DISTINCT b.STUDY_ID
                        FROM read_parquet('{str(BRIDGE_PATH)}') b
                        JOIN read_parquet('{str(USAGE_PATH)}') u ON b.USAGE_ID = u.USAGE_ID
                        JOIN read_parquet('{str(USER_GROUP_PATH)}') ug ON u.USER_ID = ug.USER_ID
                        WHERE {cross['where_str']}
                    )"""
    # Support wildcards: * becomes %, ? becomes _
    if '*' in search_lower or '?' in search_lower:
        search_like = search_lower.replace('*', '%').replace('?', '_')
    else:
        search_like = f'%{search_lower}%'
    try:
        con = duckdb.connect()
        sql = f"""SELECT DISTINCT
                      {field} as val,
                      {field} as label,
                      {field} as sort_key
                  FROM {parquet_src}
                  WHERE {field} IS NOT NULL
                    AND LOWER(CAST({field} AS VARCHAR)) LIKE '{search_like}'{where_extra}
                  ORDER BY sort_key {sort_dir}
                  LIMIT 5000"""
        result = con.execute(sql).fetchdf()
        con.close()
        options = [
            {"label": str(row["label"]).strip(),
             "value": str(row["val"]).strip()}
            for _, row in result.iterrows()
            if row["val"] is not None
        ]
        return options if options else \
               [{"label": "No matches found", "value": "__none__", "disabled": True}]
    except Exception as e:
        print(f"Search error for {field}: {e}")
        return [{"label": "Search error", "value": "__none__", "disabled": True}]
        
                   
@app.callback(
    Output("filter-search-status", "children"),
    Input("perm-field-filter-checklist", "options"),
    State("perm-field-context",          "data"),
    State("perm-filter-search",          "value"),
    prevent_initial_call=True
)
def update_search_status(options, context, search):
    if not options:
        return ""
    real = [o for o in options if o.get("value") not in
            ("__loading__", "__error__", "__hint__", "__none__")]
    if not search or not search.strip():
        return f"{len(real)} values"
    return "No matches found" if not real else f"{len(real)} matches"


@app.callback(
    Output("perm-field-filter-summary", "children", allow_duplicate=True),
    Input("perm-field-filter-checklist", "value"),
    Input("perm-filter-exclude",         "value"),
    prevent_initial_call=True
)
def update_selected_summary(values, exclude):
    return _build_filter_summary(values or [], bool(exclude))


@app.callback(
    Output({"type": "ws-field-filters",  "index": ALL}, "data"),
    Output({"type": "ws-filters",        "index": ALL}, "data", allow_duplicate=True),
    Output("perm-field-filter-modal",     "is_open",    allow_duplicate=True),
    Output("perm-field-filter-checklist", "value",      allow_duplicate=True),
    Output("perm-field-filter-summary",   "children",   allow_duplicate=True),
    Output({"type": "ws-date-formats",   "index": ALL}, "data", allow_duplicate=True),
    Input("perm-field-filter-apply", "n_clicks"),
    Input("perm-field-filter-clear", "n_clicks"),
    State("perm-field-filter-checklist", "value"),
    State("perm-filter-exclude",         "value"),
    State("perm-filter-date-fmt",        "value"),
    State("perm-field-context",          "data"),
    State({"type": "ws-field-filters",   "index": ALL}, "data"),
    State({"type": "ws-filters",         "index": ALL}, "data"),
    State({"type": "ws-date-formats",    "index": ALL}, "data"),
    prevent_initial_call=True
)
def apply_field_filter(apply_clicks, clear_clicks, checklist_values,
                       exclude, date_fmt, context, field_filters_data,
                       ws_filters_data, date_formats_data):
    triggered = ctx.triggered_id
    if triggered is None or not context or not field_filters_data:
        raise dash.exceptions.PreventUpdate
    worksheet_id = context.get("worksheet_id", "")
    field        = context.get("field", "")
    filter_key   = f"filters|{field}"
    ws_index = next((i for i, item in enumerate(ctx.states_list[4])
                     if item["id"]["index"] == worksheet_id), 0)
    date_formats = dict(date_formats_data[ws_index] if date_formats_data else {})
    num_df = len(date_formats_data) if date_formats_data else 1
    field_filters = dict(field_filters_data[ws_index])
    num_ff        = len(field_filters_data)
    num_filters   = len(ws_filters_data)
    current_fs    = list(ws_filters_data[ws_index]) if ws_filters_data else []

    if triggered == "perm-field-filter-clear":
        return (
            [dash.no_update] * num_ff,
            [dash.no_update] * num_filters,
            dash.no_update, [], "None",
            [dash.no_update] * num_df,
        )
    else:
        # Save date format
        shelf = context.get("shelf", "filters")
        if date_fmt and date_fmt != "none":
            date_formats[_fmt_key(shelf, field)] = date_fmt
        else:
            date_formats.pop(_fmt_key(shelf, field), None)
            date_formats.pop(field, None)

        if checklist_values:
            clean = [v.strip() for v in checklist_values
                if v not in ("__loading__", "__error__", "__hint__",
                             "__none__", "__cleared__")]
            if clean:
                # Save as dict with values + exclude flag
                field_filters[filter_key] = {
                    "values":  clean,
                    "exclude": bool(exclude),
                }
                if field not in current_fs:
                    current_fs.append(field)
            else:
                field_filters.pop(filter_key, None)
        else:
            field_filters.pop(filter_key, None)
        summary = _build_filter_summary(
            field_filters.get(filter_key, {}).get("values", []),
            bool(exclude)
        )
        return (
            [field_filters if i == ws_index else dash.no_update for i in range(num_ff)],
            [current_fs if i == ws_index else dash.no_update for i in range(num_filters)],
            False, dash.no_update, summary,
            [date_formats if i == ws_index else dash.no_update for i in range(num_df)],
        )


@app.callback(
    Output("perm-filter-search", "value"),
    Input("perm-field-filter-modal", "is_open"),
    Input("perm-field-context",      "data"),
    prevent_initial_call=True
)
def clear_search(is_open, context):
    return ""

def _build_cascade_where(field_filters, current_field):
    """Build DuckDB WHERE clauses from existing filters, excluding the current field."""
    cfg = load_config()
    calcs = cfg.get("global_calculations", {})

    DATE_FILTER_FIELDS = {"TABRUN_MY", "ACTION_DATE", "TABRUN_TS"}

    same_table_clauses = []
    cross_table_fields = []

    for key, saved in (field_filters or {}).items():
        if isinstance(saved, dict):
            values = saved.get("values", [])
            exclude = saved.get("exclude", False)
        else:
            values = saved
            exclude = False
        if not values:
            continue
        f = key.split("|")[-1] if "|" in key else key
        if f == current_field:
            continue

        # Skip formula calculations — they don't map to real columns
        if f in calcs:
            continue

        # Skip fields we don't recognize
        known_fields = (
            {"CLIENT_NAME", "USER_NAME", "USER_EMAIL", "ACTION_TYPE",
             "ACTION_DATE", "TABRUN_MY", "TABRUN_TS", "USER_ID", "USAGE_ID"}
            | _STUDY_FILTER_FIELDS | _USER_GROUP_FILTER_FIELDS
        )
        if f not in known_fields:
            continue

        escaped = [v.strip().replace("'", "''") for v in values]

        # Date fields need LIKE matching, not IN
        if f in DATE_FILTER_FIELDS:
            parts = []
            for v in escaped:
                q_match = re.match(r'^Q(\d)\s+(\d{4})$', v, re.IGNORECASE)
                if not q_match:
                    q_match = re.match(r'^(\d{4})\s+Q(\d)$', v, re.IGNORECASE)
                    if q_match:
                        yr, qtr = q_match.group(1), q_match.group(2)
                    else:
                        yr, qtr = None, None
                else:
                    qtr, yr = q_match.group(1), q_match.group(2)

                if yr and qtr:
                    if exclude:
                        parts.append(f"NOT (YEAR({f}) = {yr} AND QUARTER({f}) = {int(qtr)})")
                    else:
                        parts.append(f"(YEAR({f}) = {yr} AND QUARTER({f}) = {int(qtr)})")
                else:
                    if exclude:
                        parts.append(f"{f}::VARCHAR NOT LIKE '%{v}%'")
                    else:
                        parts.append(f"{f}::VARCHAR LIKE '%{v}%'")

            joiner = " AND " if exclude else " OR "
            where_str = f"({joiner.join(parts)})"
        else:
            vals = ", ".join([f"'{v}'" for v in escaped])
            not_kw = "NOT " if exclude else ""
            where_str = f"{f} {not_kw}IN ({vals})"

        # Determine which table this filter field lives in
        if f in _STUDY_FILTER_FIELDS:
            filter_table = "study"
        elif f in _USER_GROUP_FILTER_FIELDS:
            filter_table = "user_group"
        else:
            filter_table = "usage"

        # Determine which table the current field lives in
        if current_field in _STUDY_FILTER_FIELDS:
            current_table = "study"
        elif current_field in _USER_GROUP_FILTER_FIELDS:
            current_table = "user_group"
        else:
            current_table = "usage"

        if filter_table == current_table:
            same_table_clauses.append(where_str)
        else:
            cross_table_fields.append({
                "where_str": where_str,
                "filter_table": filter_table, "current_table": current_table
            })

    return same_table_clauses, cross_table_fields


# ─────────────────────────────────────────────
# Run query
# ─────────────────────────────────────────────

@app.callback(
    Output({"type": "data-table-container", "index": ALL}, "children"),
    Output("ws-last-run-state", "data", allow_duplicate=True),
    Input({"type": "run-query-btn",          "index": ALL}, "n_clicks"),
    State({"type": "ws-rows",                "index": ALL}, "data"),
    State({"type": "ws-cols",                "index": ALL}, "data"),
    State({"type": "ws-field-filters",       "index": ALL}, "data"),
    State({"type": "ws-date-formats",        "index": ALL}, "data"),
    State({"type": "ws-measure",             "index": ALL}, "data"),
    State("global-calculations", "data"),
    State("ws-settings-store",   "data"),
    State("ws-last-run-state",   "data"),
    State({"type": "ws-filters", "index": ALL}, "data"),
    prevent_initial_call=True
)
def run_worksheet_query(n_clicks, rows_data, cols_data, field_filters_data,
                        date_formats_data, measure_data, global_calcs,
                        ws_settings, last_run, ws_filters_data):
    
    triggered = ctx.triggered_id
    if triggered is None:
        raise dash.exceptions.PreventUpdate
    ti = next((i for i, item in enumerate(ctx.inputs_list[0])
    if item["id"]["index"] == triggered["index"]), 0)

    rows          = rows_data[ti]          if rows_data          else []
    cols          = cols_data[ti]          if cols_data          else []
    field_filters = field_filters_data[ti] if field_filters_data else {}
    date_formats  = date_formats_data[ti]  if date_formats_data  else {}
    measure       = measure_data[ti]       if measure_data       else "COUNTD_USAGE_ID"
    filters       = ws_filters_data[ti]    if ws_filters_data    else []
    all_fields    = rows + [f for f in cols if f not in rows]

    if not all_fields:
        result = html.P("Add at least one field to Rows or Columns.",
                       className="text-warning p-3")
        return [result if i == ti else dash.no_update
                for i in range(len(n_clicks))], dash.no_update

    # Separate formula calcs from regular fields
    formula_calcs = {}
    formula_extra_fields = set()
    if isinstance(global_calcs, dict):
        for f in all_fields:
            calc_name = f[5:] if f.startswith("calc_") else f
            defn = global_calcs.get(calc_name, global_calcs.get(f, {}))
            if isinstance(defn, dict) and defn.get("type") == "formula":
                formula_calcs[f] = defn.get("formula", "")
                formula_str = defn.get("formula", "")
                formula_upper = formula_str.upper()
                for ref_field in FORMULA_FIELDS:
                    if ref_field.upper() in formula_upper and ref_field not in all_fields:
                        formula_extra_fields.add(ref_field)

    duck_fields = [f for f in all_fields if f not in formula_calcs]
    for ef in formula_extra_fields:
        if ef not in duck_fields:
            duck_fields.append(ef)

    if not duck_fields:
        result = html.P("Add at least one non-formula field to Rows or Columns.",
                       className="text-warning p-3")
        return [result if i == ti else dash.no_update
                for i in range(len(n_clicks))], dash.no_update

    needs_qmnem     = False
    duck_where      = []
    query_date_formats = {}
    use_duckdb      = False

    if extract_available():
        for f in duck_fields:
            if f in formula_extra_fields:
                continue
            fmt = _get_fmt(date_formats, "rows" if f in rows else "cols", f)
            if fmt != "none":
                query_date_formats[f] = fmt
        duck_where = build_duck_where(field_filters, date_formats)
        sql = build_duckdb_query(
            duck_fields, duck_where, date_formats=query_date_formats, measure=measure,
            global_calcs=global_calcs if isinstance(global_calcs, dict) else {}
        )
        use_duckdb = True
    else:
        where_clauses, params = [], {}
        for k, values in (field_filters or {}).items():
            if values:
                f = k.split("|")[-1] if "|" in k else k
                phs = ", ".join([f":ff_{f}_{i}" for i in range(len(values))])
                where_clauses.append(f"{get_field_expr(f)} IN ({phs})")
                for i, v in enumerate(values):
                    params[f"ff_{f}_{i}"] = v
        sql, needs_qmnem = build_query(all_fields, where_clauses, params)

    warning = None
    if needs_qmnem:
        warning = dbc.Alert("⚠️ QMNEM join — may be slow.", color="warning",
                           className="mb-2 py-1", style={"fontSize": "12px"})

    ws_key = triggered["index"]
    ws_cfg = (ws_settings or {}).get(ws_key, {})
    results_dir = get_base_dir() / "data" / "results"
    results_dir.mkdir(parents=True, exist_ok=True)

    try:
        if use_duckdb:
            df = run_extract_query(sql)

            # ── Apply FIXED formulas BEFORE pivot (raw data has all columns) ──
            if formula_calcs and global_calcs and isinstance(global_calcs, dict):
                for calc_name, defn in global_calcs.items():
                    if isinstance(defn, dict) and defn.get("type") == "formula":
                        if calc_name not in all_fields and f"calc_{calc_name}" not in all_fields:
                            continue
                        try:
                            df[calc_name] = apply_calculation(defn["formula"], df)
                            if f"calc_{calc_name}" in all_fields and calc_name not in all_fields:
                                df[f"calc_{calc_name}"] = df[calc_name]
                                df.drop(columns=[calc_name], inplace=True)
                        except Exception as e:
                            print(f"Pre-pivot calc error {calc_name}: {e}")

                # Drop extra fields fetched only for formulas
                for ef in formula_extra_fields:
                    if ef in df.columns and ef not in rows and ef not in cols:
                        df.drop(columns=[ef], inplace=True)

            # Save results AFTER formulas are applied
            try:
                df.to_parquet(results_dir / f"{ws_key}.parquet", index=False)
                with open(results_dir / f"{ws_key}.meta.json", "w") as f:
                    json.dump({"rows": rows, "cols": cols,
                               "date_formats": date_formats, "measure": measure}, f)
            except Exception as e:
                print(f"Could not save result: {e}")

            if cols and "Count" in df.columns:
                if not rows:
                    # No row fields — pivot columns into a single row
                    pivot = df.set_index(cols[0])["Count"]
                    df = pd.DataFrame([pivot.values], columns=pivot.index)
                    if cols[0] in DATE_FIELDS:
                        try:
                            sorted_cols = sorted(df.columns,
                                                 key=lambda x: pd.to_datetime(x, format='%b %Y'))
                        except Exception:
                            sorted_cols = list(df.columns)
                        df = df[sorted_cols]
                    df["Grand Total"] = df.sum(axis=1)
                else:
                    # ── Run summary query for correct grand totals ────────────
                    FANOUT_FIELDS = {"LONG_NAME", "EXT_STUDY_ID", "STUDYID",
                                     "STUDYYEAR", "GROUP_NAME", "USER_EMAIL"}
                    has_fanout = any(f in FANOUT_FIELDS for f in rows)
                    df_summary = None
                    if has_fanout:
                        summary_row_fields = [f for f in rows
                                              if f not in FANOUT_FIELDS and f not in formula_calcs]
                        summary_fields = summary_row_fields + [
                            f for f in cols if f not in summary_row_fields
                        ]
                        if summary_fields:
                            try:
                                summary_date_fmts = {
                                    f: query_date_formats[f]
                                    for f in summary_fields
                                    if f in query_date_formats
                                }
                                summary_sql = build_duckdb_query(
                                    summary_fields, duck_where,
                                    date_formats=summary_date_fmts,
                                    measure=measure,
                                    global_calcs=global_calcs
                                    if isinstance(global_calcs, dict) else {}
                                )
                                df_summary = run_extract_query(summary_sql)
                            except Exception as e:
                                print(f"Summary query error: {e}")
                                df_summary = None

                    # Save summary for restore
                    if df_summary is not None:
                        try:
                            df_summary.to_parquet(
                                results_dir / f"{ws_key}.summary.parquet", index=False)
                        except Exception:
                            pass

                    df = apply_pivot(df, rows, cols,
                                     col_total=ws_cfg.get("col_grand_total", "last"),
                                     row_total=ws_cfg.get("row_grand_total", "first"),
                                     df_summary=df_summary)
        else:
            df = run_query(sql, params)

        # Post-pivot formulas (non-FIXED, e.g. math on Count columns)
        if global_calcs and isinstance(global_calcs, dict):
            for calc_name, defn in global_calcs.items():
                if isinstance(defn, dict) and defn.get("type") == "formula":
                    if calc_name in all_fields or f"calc_{calc_name}" in all_fields:
                        if calc_name not in df.columns and f"calc_{calc_name}" not in df.columns:
                            try:
                                df[calc_name] = apply_calculation(defn["formula"], df)
                            except Exception as e:
                                print(f"Post-pivot calc error {calc_name}: {e}")

        display_df = apply_row_blanking(df, rows)
        table      = build_html_table(df, display_df=display_df, rows=rows)
        _save_ws_state(triggered["index"], rows, cols, filters,
                       field_filters, date_formats, measure)

        result = html.Div([
            warning,
            html.Div([
                html.Span(f"{len(display_df):,} rows returned",
                         className="text-muted", style={"fontSize": "11px"}),
                html.Div([
                    dbc.Button(
                        [html.I(className="bi bi-file-earmark-excel me-1"), "Spreadsheet"],
                        id={"type": "export-btn", "index": triggered["index"]},
                        color="success", size="sm", outline=True,
                        style={"fontSize": "11px", "padding": "1px 8px"}
                    ),
                    dbc.Button(
                        [html.I(className="bi bi-file-earmark-excel me-1"), "Crosstab"],
                        id={"type": "export-crosstab-btn", "index": triggered["index"]},
                        color="primary", size="sm", outline=True,
                        style={"fontSize": "11px", "padding": "1px 8px"}
                    ),
                ], className="d-flex gap-1"),
            ], className="d-flex align-items-center justify-content-between px-2 py-1",
               style={"borderBottom": "1px solid #eee"}),
            table
        ])
    except Exception as e:
        print(f"Query error: {str(e)}")
        result = html.P(f"Query error: {str(e)}", className="text-danger p-3")

    last_run_update = dict(last_run or {})
    if not isinstance(result, html.P) or "Query error" not in (result.children or ""):
        last_run_update[triggered["index"]] = {
            "rows": rows, "cols": cols,
            "field_filters": field_filters, "measure": measure,
        }
    return ([result if i == ti else dash.no_update for i in range(len(n_clicks))],
            last_run_update if last_run_update != (last_run or {}) else dash.no_update)

@app.callback(
    Output({"type": "data-table-container", "index": ALL}, "children",
           allow_duplicate=True),
    Output("ws-last-run-state", "data", allow_duplicate=True),
    Input("worksheet-content", "children"),
    prevent_initial_call='initial_duplicate'
)
def restore_saved_results(children):
    if not children:
        raise dash.exceptions.PreventUpdate
    cfg = load_config()
    worksheets = cfg.get("worksheets", [])
    if not worksheets:
        raise dash.exceptions.PreventUpdate
    ws_settings  = cfg.get("ws_settings", {})
    ws_state_all = cfg.get("ws_state", {})
    results      = []
    last_run_out = {}
    for w in worksheets:
        ws_key       = w.replace(" ", "_")
        results_dir  = get_base_dir() / "data" / "results"
        parquet_path = results_dir / f"{ws_key}.parquet"
        meta_path    = results_dir / f"{ws_key}.meta.json"
        if not parquet_path.exists() or not meta_path.exists():
            results.append(dash.no_update)
            continue
        try:
            df = pd.read_parquet(parquet_path)
            with open(meta_path) as f:
                meta = json.load(f)
            rows = meta.get("rows", [])
            cols = meta.get("cols", [])

            if cols and rows and "Count" in df.columns:
                ws_cfg = ws_settings.get(ws_key, {})

                df_summary = None
                summary_path = results_dir / f"{ws_key}.summary.parquet"
                if summary_path.exists():
                    try:
                        df_summary = pd.read_parquet(summary_path)
                    except Exception:
                        pass

                df = apply_pivot(df, rows, cols,
                                 col_total=ws_cfg.get("col_grand_total", "last"),
                                 row_total=ws_cfg.get("row_grand_total", "first"),
                                 df_summary=df_summary)

            # Skip large results — user can hit Run Query
            if len(df) > 1000:
                ws_state = ws_state_all.get(ws_key, {})
                last_run_out[ws_key] = {
                    "rows":          ws_state.get("rows", []),
                    "cols":          ws_state.get("cols", []),
                    "field_filters": ws_state.get("field_filters", {}),
                    "measure":       ws_state.get("measure", "COUNTD_USAGE_ID"),
                }
                results.append(html.Div([
                    html.Div([
                        html.I(className="bi bi-table",
                               style={"fontSize": "24px", "marginBottom": "8px",
                                      "color": "#6c757d"}),
                        html.Div(f"{len(df):,} rows — too large to auto-load.",
                                 style={"fontSize": "13px", "fontWeight": "bold",
                                        "color": "#555"}),
                        html.Div("Click Run Query to load results.",
                                 style={"fontSize": "11px", "color": "#888",
                                        "marginTop": "4px"}),
                    ], style={"textAlign": "center", "padding": "60px 30px"})
                ]))
                continue

            display_df = apply_row_blanking(df, rows)
            table      = build_html_table(df, display_df=display_df, rows=rows)
            ws_state   = ws_state_all.get(ws_key, {})
            last_run_out[ws_key] = {
                "rows":          ws_state.get("rows", []),
                "cols":          ws_state.get("cols", []),
                "field_filters": ws_state.get("field_filters", {}),
                "measure":       ws_state.get("measure", "COUNTD_USAGE_ID"),
            }
            results.append(html.Div([
                html.Div([
                    html.Span(f"{len(display_df):,} rows — saved result",
                             className="text-muted fst-italic",
                             style={"fontSize": "11px"}),
                    html.Div([
                        dbc.Button(
                            [html.I(className="bi bi-file-earmark-excel me-1"), "Spreadsheet"],
                            id={"type": "export-btn", "index": ws_key},
                            color="success", size="sm", outline=True,
                            style={"fontSize": "11px", "padding": "1px 8px"}
                        ),
                        dbc.Button(
                            [html.I(className="bi bi-file-earmark-excel me-1"), "Crosstab"],
                            id={"type": "export-crosstab-btn", "index": ws_key},
                            color="primary", size="sm", outline=True,
                            style={"fontSize": "11px", "padding": "1px 8px"}
                        ),
                    ], className="d-flex gap-1"),
                ], className="d-flex align-items-center justify-content-between px-2 py-1",
                   style={"borderBottom": "1px solid #eee"}),
                table
            ], style={"display": "flex", "flexDirection": "column",
                      "height": "100%", "minHeight": "0", "overflow": "hidden"}))
        except Exception as e:
            print(f"Could not restore {ws_key}: {e}")
            results.append(dash.no_update)
    return results, last_run_out


# ─────────────────────────────────────────────
# Export
# ─────────────────────────────────────────────

@app.callback(
    Output("download-data", "data"),
    Input({"type": "export-btn",       "index": ALL}, "n_clicks"),
    State({"type": "ws-rows",          "index": ALL}, "data"),
    State({"type": "ws-cols",          "index": ALL}, "data"),
    State({"type": "ws-field-filters", "index": ALL}, "data"),
    State({"type": "ws-date-formats",  "index": ALL}, "data"),
    State({"type": "ws-measure",       "index": ALL}, "data"),
    State("global-calculations", "data"),
    State("ws-settings-store",   "data"),
    State("worksheet-tabs",      "value"),
    prevent_initial_call=True
)
def export_data(n_clicks, rows_data, cols_data, field_filters_data,
                date_formats_data, measure_data, global_calcs, ws_settings, active_tab):
    if not any(n for n in n_clicks if n):
        raise dash.exceptions.PreventUpdate
    triggered = ctx.triggered_id
    ti = next((i for i, item in enumerate(ctx.inputs_list[0])
               if item["id"] == triggered), 0)
    field_filters = field_filters_data[ti] if field_filters_data else {}
    date_formats  = date_formats_data[ti]  if date_formats_data  else {}
    measure       = measure_data[ti]       if measure_data       else "COUNTD_USAGE_ID"

    try:
        ws_key       = triggered["index"] if isinstance(triggered, dict) \
                       else active_tab.replace(" ", "_")
        results_dir  = get_base_dir() / "data" / "results"
        parquet_path = results_dir / f"{ws_key}.parquet"
        meta_path    = results_dir / f"{ws_key}.meta.json"

        rows = rows_data[ti] if rows_data else []
        cols = cols_data[ti] if cols_data else []
        if meta_path.exists():
            with open(meta_path) as mf:
                meta = json.load(mf)
            rows = meta.get("rows", rows)
            cols = meta.get("cols", cols)

        all_fields = rows + [f for f in cols if f not in rows]
        if not all_fields:
            raise dash.exceptions.PreventUpdate

        if parquet_path.exists():
            df = pd.read_parquet(parquet_path)
        else:
            query_date_formats = {}
            for f in all_fields:
                fmt = _get_fmt(date_formats, "rows" if f in rows else "cols", f)
                if fmt != "none":
                    query_date_formats[f] = fmt
            duck_where = build_duck_where(field_filters, date_formats)
            sql = build_duckdb_query(
                all_fields, duck_where, date_formats=query_date_formats, measure=measure,
                global_calcs=global_calcs if isinstance(global_calcs, dict) else {}
            )
            df = run_extract_query(sql)

        if cols and "Count" in df.columns:
            if not rows:
                # No row fields — pivot columns into a single row
                pivot = df.set_index(cols[0])["Count"]
                df = pd.DataFrame([pivot.values], columns=pivot.index)
                if cols[0] in DATE_FIELDS:
                    try:
                        sorted_cols = sorted(df.columns, key=lambda x: pd.to_datetime(x, format='%b %Y'))
                    except Exception:
                        sorted_cols = list(df.columns)
                    df = df[sorted_cols]
                df["Grand Total"] = df.sum(axis=1)
            else:
                ws_cfg = (ws_settings or {}).get(ws_key, {})

                df_summary = None
                summary_path = results_dir / f"{ws_key}.summary.parquet"
                if summary_path.exists():
                    try:
                        df_summary = pd.read_parquet(summary_path)
                    except Exception:
                        pass

                df = apply_pivot(df, rows, cols,
                                col_total=ws_cfg.get("col_grand_total", "last"),
                                row_total=ws_cfg.get("row_grand_total", "first"),
                                df_summary=df_summary)

        import io
        output = io.BytesIO()
        df.to_excel(output, index=False)
        output.seek(0)
        return dcc.send_bytes(
            output.getvalue(),
            filename=f"{active_tab.replace(' ', '_')}.xlsx"
        )

    except dash.exceptions.PreventUpdate:
        raise
    except Exception as e:
        print(f"Export error: {e}")
        raise dash.exceptions.PreventUpdate


@app.callback(
    Output("download-crosstab", "data"),
    Input({"type": "export-crosstab-btn", "index": ALL}, "n_clicks"),
    State({"type": "ws-rows",          "index": ALL}, "data"),
    State({"type": "ws-cols",          "index": ALL}, "data"),
    State({"type": "ws-field-filters", "index": ALL}, "data"),
    State({"type": "ws-date-formats",  "index": ALL}, "data"),
    State({"type": "ws-measure",       "index": ALL}, "data"),
    State("global-calculations", "data"),
    State("ws-settings-store",   "data"),
    State("worksheet-tabs",      "value"),
    prevent_initial_call=True
)
def export_crosstab(n_clicks, rows_data, cols_data, field_filters_data,
                    date_formats_data, measure_data, global_calcs, ws_settings, active_tab):
    if not any(n for n in n_clicks if n):
        raise dash.exceptions.PreventUpdate
    triggered = ctx.triggered_id
    if triggered is None:
        raise dash.exceptions.PreventUpdate
    ti = next((i for i, item in enumerate(ctx.inputs_list[0])
               if item["id"] == triggered), 0)
    field_filters = field_filters_data[ti] if field_filters_data else {}
    date_formats  = date_formats_data[ti]  if date_formats_data  else {}
    measure       = measure_data[ti]       if measure_data       else "COUNTD_USAGE_ID"

    try:
        ws_key      = triggered["index"] if isinstance(triggered, dict) \
                      else active_tab.replace(" ", "_")
        results_dir = get_base_dir() / "data" / "results"
        parquet_path = results_dir / f"{ws_key}.parquet"
        meta_path    = results_dir / f"{ws_key}.meta.json"

        rows = rows_data[ti] if rows_data else []
        cols = cols_data[ti] if cols_data else []
        if meta_path.exists():
            with open(meta_path) as mf:
                meta = json.load(mf)
            rows = meta.get("rows", rows)
            cols = meta.get("cols", cols)

        all_fields = rows + [f for f in cols if f not in rows]
        if not all_fields:
            raise dash.exceptions.PreventUpdate

        if parquet_path.exists():
            df = pd.read_parquet(parquet_path)
        else:
            query_date_formats = {}
            for f in all_fields:
                fmt = _get_fmt(date_formats, "rows" if f in rows else "cols", f)
                if fmt != "none":
                    query_date_formats[f] = fmt
            duck_where = build_duck_where(field_filters, date_formats)
            sql = build_duckdb_query(
                all_fields, duck_where, date_formats=query_date_formats, measure=measure,
                global_calcs=global_calcs if isinstance(global_calcs, dict) else {}
            )
            df = run_extract_query(sql)

        if cols and "Count" in df.columns:
            if not rows:
                # No row fields — pivot columns into a single row
                pivot = df.set_index(cols[0])["Count"]
                df = pd.DataFrame([pivot.values], columns=pivot.index)
                if cols[0] in DATE_FIELDS:
                    try:
                        sorted_cols = sorted(df.columns, key=lambda x: pd.to_datetime(x, format='%b %Y'))
                    except Exception:
                        sorted_cols = list(df.columns)
                    df = df[sorted_cols]
                df["Grand Total"] = df.sum(axis=1)
            else:
                ws_cfg = (ws_settings or {}).get(ws_key, {})

                df_summary = None
                summary_path = results_dir / f"{ws_key}.summary.parquet"
                if summary_path.exists():
                    try:
                        df_summary = pd.read_parquet(summary_path)
                    except Exception:
                        pass

                df = apply_pivot(df, rows, cols,
                                col_total=ws_cfg.get("col_grand_total", "last"),
                                row_total=ws_cfg.get("row_grand_total", "first"),
                                df_summary=df_summary)
        display_df = apply_row_blanking(df, rows)

        import io
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = active_tab[:31]
        header_fill  = PatternFill("solid", fgColor="0F1F3D")
        measure_fill = PatternFill("solid", fgColor="1E3A5F")
        gt_fill      = PatternFill("solid", fgColor="2D6A4F")
        gt_row_fill  = PatternFill("solid", fgColor="FFF3CD")
        white_font   = Font(color="FFFFFF", bold=True, size=10)
        normal_font  = Font(size=10)
        bold_font    = Font(bold=True, size=10)
        center_align = Alignment(horizontal="center", vertical="top", wrap_text=True)
        left_align   = Alignment(horizontal="left",   vertical="top", wrap_text=True)
        thin         = Side(style="thin", color="ADB5BD")
        thin_border  = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx, col_name in enumerate(display_df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = (gt_fill if col_name == "Grand Total"
                         else measure_fill if col_name not in rows
                         else header_fill)
            cell.font = white_font
            cell.alignment = center_align
            cell.border = thin_border

        for row_idx, (_, row) in enumerate(display_df.iterrows(), 2):
            is_gt = (str(df.iloc[row_idx - 2][df.columns[0]]) == "Grand Total"
                     if row_idx - 2 < len(df) else False)
            for col_idx, col_name in enumerate(display_df.columns, 1):
                val = row[col_name]
                if col_name not in rows:
                    val = df.iloc[row_idx - 2][col_name] if row_idx - 2 < len(df) else val
                cell = ws.cell(row=row_idx, column=col_idx,
                               value=None if str(val) == "" else val)
                cell.font      = bold_font if is_gt else normal_font
                cell.alignment = left_align if col_name in rows else center_align
                cell.border    = thin_border
                if is_gt:
                    cell.fill = gt_row_fill

        for col_idx, col_name in enumerate(display_df.columns, 1):
            if col_name not in rows:
                continue
            merge_start = None
            for row_idx in range(2, len(display_df) + 2):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val is None or cell_val == "":
                    if merge_start is None:
                        merge_start = row_idx - 1
                else:
                    if merge_start is not None and row_idx - 1 > merge_start:
                        ws.merge_cells(start_row=merge_start, start_column=col_idx,
                                       end_row=row_idx - 1, end_column=col_idx)
                        ws.cell(row=merge_start, column=col_idx).alignment = Alignment(
                            horizontal="left", vertical="top", wrap_text=True)
                    merge_start = None
            if merge_start is not None and len(display_df) + 1 > merge_start:
                ws.merge_cells(start_row=merge_start, start_column=col_idx,
                               end_row=len(display_df) + 1, end_column=col_idx)

        for col_idx, col_name in enumerate(display_df.columns, 1):
            max_len = max(len(str(col_name)),
                         *[len(str(ws.cell(row=r, column=col_idx).value or ""))
                           for r in range(2, len(display_df) + 2)])
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col_idx)
            ].width = min(max_len + 2, 40)

        ws.freeze_panes = "A2"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return dcc.send_bytes(
            output.getvalue(),
            filename=f"{active_tab.replace(' ', '_')}_crosstab.xlsx"
        )

    except dash.exceptions.PreventUpdate:
        raise
    except Exception as e:
        print(f"Crosstab export error: {e}")
        import traceback; traceback.print_exc()
        raise dash.exceptions.PreventUpdate


# ─────────────────────────────────────────────
# Settings
# ─────────────────────────────────────────────

@app.callback(
    Output("settings-modal",      "is_open"),
    Output("settings-modal-body", "children"),
    Input("settings-btn",         "n_clicks"),
    State("settings-modal",       "is_open"),
    prevent_initial_call=True
)
def toggle_settings(n_clicks, is_open):
    if n_clicks:
        return not is_open, build_settings_layout()
    return is_open, dash.no_update


@app.callback(
    Output("refresh-extract-btn", "n_clicks"),
    Input("settings-refresh-btn", "n_clicks"),
    State("refresh-extract-btn",  "n_clicks"),
    prevent_initial_call=True
)
def relay_refresh(n, current):
    if not n:
        raise dash.exceptions.PreventUpdate
    return (current or 0) + 1


@app.callback(
    Output("cancel-extract-btn", "n_clicks"),
    Input("settings-cancel-btn", "n_clicks"),
    State("cancel-extract-btn",  "n_clicks"),
    prevent_initial_call=True
)
def relay_cancel(n, current):
    if not n:
        raise dash.exceptions.PreventUpdate
    return (current or 0) + 1


@app.callback(
    Output("extract-progress-display", "children"),
    Input("extract-interval",          "n_intervals"),
    Input("refresh-extract-btn",       "n_clicks"),
    prevent_initial_call=True
)
def update_settings_progress(n_intervals, n_clicks):
    if not extract_messages:
        raise dash.exceptions.PreventUpdate
    latest = extract_messages[-1]
    color  = ("success" if latest.startswith("✅") else
              "danger"  if latest.startswith("❌") else "info")
    return dbc.Alert(latest, color=color, className="mb-0 py-1 mt-2",
                    style={"fontSize": "12px"})


@app.callback(
    Output("settings-cancel-btn", "style"),
    Input("extract-interval",     "n_intervals"),
    Input("refresh-extract-btn",  "n_clicks"),
    prevent_initial_call=True
)
def toggle_settings_cancel(n_intervals, n_clicks):
    if not extract_messages:
        return {"display": "none"}
    latest  = extract_messages[-1]
    is_done = latest.startswith("✅") or latest.startswith("❌")
    return {"display": "none"} if is_done else {"display": "inline-block"}


@app.callback(
    Output("cfg-save-msg", "children"),
    Input("cfg-save-btn",    "n_clicks"),
    State("cfg-host",        "value"),
    State("cfg-port",        "value"),
    State("cfg-dbname",      "value"),
    State("cfg-user",        "value"),
    State("cfg-password",    "value"),
    State("cfg-upload-path", "value"),
    prevent_initial_call=True
)
def save_config_callback(n_clicks, host, port, dbname, user, password, upload_path):
    if not n_clicks:
        raise dash.exceptions.PreventUpdate
    cfg = load_config()
    cfg.update({
        "DB_HOST":           host or "",
        "DB_PORT":           port or "3306",
        "DB_NAME":           dbname or "",
        "DB_USER":           user or "",
        "DB_PASSWORD":       password or "",
        "MYSQL_UPLOAD_PATH": upload_path or "",
    })
    save_config(cfg)
    return dbc.Alert("✅ Configuration saved!", color="success",
                    className="mb-0 py-1", style={"fontSize": "12px"})


# ─────────────────────────────────────────────
# Extract progress
# ─────────────────────────────────────────────

extract_messages   = []
extract_pct        = 0
extract_start_time = None


@app.callback(
    Output("extract-progress",                "children"),
    Output("extract-progress-bar-container",  "style"),
    Output("extract-progress-bar",            "value"),
    Output("extract-progress-bar",            "label"),
    Output("extract-interval",                "disabled"),
    Output("refresh-extract-btn",             "disabled"),
    Output("extract-progress-persistent",     "style"),
    Output("extract-progress-persistent-msg", "children"),
    Output("extract-progress-persistent-bar", "value"),
    Output("cancel-extract-btn",              "style"),
    Input("refresh-extract-btn", "n_clicks"),
    Input("extract-interval",    "n_intervals"),
    prevent_initial_call=True
)
def handle_extract(n_clicks, n_intervals):
    global extract_messages, extract_pct, extract_start_time
    triggered = ctx.triggered_id

    if triggered == "refresh-extract-btn" and n_clicks:
        extract_messages   = ["Starting extract..."]
        extract_pct        = 0
        extract_start_time = time.time()

        def run_extract():
            global extract_messages, extract_pct
            def progress(msg, pct):
                extract_messages.append(msg)
                extract_pct = pct
            build_extract(progress_callback=progress)
            _cached_filter_options.cache_clear()

        threading.Thread(target=run_extract, daemon=True).start()
        visible = {"display": "block"}
        return (
            html.Div("Starting...", className="text-muted", style={"fontSize": "12px"}),
            visible, 0, "0%", False, True,
            visible, "Starting extract...", 0,
            {"display": "inline-block"},
        )

    if triggered == "extract-interval":
        if not extract_messages:
            raise dash.exceptions.PreventUpdate
        latest  = extract_messages[-1]
        is_done = latest.startswith("✅") or latest.startswith("❌")
        pct     = extract_pct
        progress_display = html.Div(latest, style={
            "fontSize": "12px",
            "color": ("green" if latest.startswith("✅") else
                      "red"   if latest.startswith("❌") else "inherit")
        })
        return (
            progress_display,
            {"display": "none"} if is_done else {"display": "block"},
            pct, f"{pct}%", is_done, is_done,
            {"display": "none"} if is_done else {"display": "block"},
            latest, pct,
            {"display": "none"} if is_done else {"display": "inline-block"},
        )
    raise dash.exceptions.PreventUpdate


@app.callback(
    Output("extract-clock",          "style"),
    Output("extract-clock-display",  "children"),
    Output("extract-clock-interval", "disabled"),
    Input("extract-clock-interval",  "n_intervals"),
    Input("refresh-extract-btn",     "n_clicks"),
    prevent_initial_call=True
)
def update_clock(n_intervals, n_clicks):
    triggered = ctx.triggered_id
    if triggered == "refresh-extract-btn" and n_clicks:
        return {"display": "block"}, "⏱ Extract starting...", False
    if extract_start_time:
        secs = int(time.time() - extract_start_time)
        mins, secs = divmod(secs, 60)
        is_done = any(m.startswith("✅") or m.startswith("❌")
                     for m in extract_messages)
        if is_done:
            return {"display": "none"}, "", True
        return {"display": "block"}, f"⏱ Extract running — {mins}m {secs}s", False
    raise dash.exceptions.PreventUpdate


@app.callback(
    Output("extract-interval", "disabled", allow_duplicate=True),
    Input("cancel-extract-btn", "n_clicks"),
    prevent_initial_call=True
)
def cancel_extract(n_clicks):
    if n_clicks:
        from data.extract import request_cancel
        request_cancel()
    raise dash.exceptions.PreventUpdate


# ─────────────────────────────────────────────
# Calculation modal
# ─────────────────────────────────────────────

@app.callback(
    Output("calc-modal",          "is_open"),
    Output("calc-name-input",     "value"),
    Output("calc-type-select",    "value"),
    Output("calc-agg-func",       "value"),
    Output("calc-agg-field",      "value"),
    Output("calc-formula-input",  "value"),
    Output("calc-lod-dim",        "value"),
    Output("calc-lod-agg",        "value"),
    Output("calc-lod-field",      "value"),
    Output("calc-validation-msg", "children"),
    Input({"type": "open-calc-modal", "index": ALL}, "n_clicks"),
    Input("calc-cancel-btn", "n_clicks"),
    Input("calc-save-btn",   "n_clicks"),
    State("calc-modal",      "is_open"),
    prevent_initial_call=True
)
def toggle_calc_modal(open_clicks, cancel, save, is_open):
    triggered = ctx.triggered_id
    blank     = (False, "", "aggregate", "COUNTD", "USAGE_ID", "",
                 "EXT_STUDY_ID", "MAX", "TABRUN_TS", "")
    if triggered == "calc-cancel-btn":
        return blank
    if triggered == "calc-save-btn":
        return (dash.no_update,) + tuple([dash.no_update] * 9)
    if isinstance(triggered, dict) and triggered.get("type") == "open-calc-modal":
        if any(n for n in open_clicks if n):
            return (True,) + blank[1:]
    return (is_open,) + tuple([dash.no_update] * 9)


@app.callback(
    Output("calc-aggregate-section", "style"),
    Output("calc-formula-section",   "style"),
    Output("calc-lod-section",       "style"),
    Input("calc-type-select", "value"),
)
def toggle_calc_sections(calc_type):
    if calc_type == "aggregate":
        return {"display": "block"}, {"display": "none"}, {"display": "none"}
    elif calc_type == "formula":
        return {"display": "none"}, {"display": "block"}, {"display": "none"}
    else:
        return {"display": "none"}, {"display": "none"}, {"display": "block"}


@app.callback(
    Output("calc-lod-preview", "children"),
    Input("calc-lod-dim",   "value"),
    Input("calc-lod-agg",   "value"),
    Input("calc-lod-field", "value"),
)
def preview_lod(dim, agg, field):
    if not dim or not agg or not field:
        return ""
    return f"{agg}({field}) OVER (PARTITION BY {dim})"


@app.callback(
    Output("calc-modal",          "is_open",  allow_duplicate=True),
    Output("calc-name-input",     "value",    allow_duplicate=True),
    Output("calc-type-select",    "value",    allow_duplicate=True),
    Output("calc-agg-func",       "value",    allow_duplicate=True),
    Output("calc-agg-field",      "value",    allow_duplicate=True),
    Output("calc-formula-input",  "value",    allow_duplicate=True),
    Output("calc-lod-dim",        "value",    allow_duplicate=True),
    Output("calc-lod-agg",        "value",    allow_duplicate=True),
    Output("calc-lod-field",      "value",    allow_duplicate=True),
    Output("calc-validation-msg", "children", allow_duplicate=True),
    Output("editing-calc-name",   "data"),
    Input({"type": "edit-calc-btn", "index": ALL}, "n_clicks"),
    State("global-calculations", "data"),
    prevent_initial_call=True
)
def edit_calc(n_clicks, global_calcs):
    if not any(n for n in n_clicks if n):
        raise dash.exceptions.PreventUpdate
    triggered = ctx.triggered_id
    if triggered is None:
        raise dash.exceptions.PreventUpdate
    name      = triggered["index"]
    defn      = (global_calcs or {}).get(name, {})
    calc_type = defn.get("type", "aggregate")
    return (True, name, calc_type,
            defn.get("func",    "COUNTD"),
            defn.get("field",   "USAGE_ID"),
            defn.get("formula", ""),
            defn.get("dim",     "EXT_STUDY_ID"),
            defn.get("agg",     "MAX"),
            defn.get("field",   "TABRUN_TS") if calc_type == "fixed_lod" else "TABRUN_TS",
            "", name)


@app.callback(
    Output("global-calculations",                    "data", allow_duplicate=True),
    Output({"type": "measure-select", "index": ALL}, "options", allow_duplicate=True),
    Output({"type": "left-panel",     "index": ALL}, "children", allow_duplicate=True),
    Input({"type": "delete-calc-btn", "index": ALL}, "n_clicks"),
    State("global-calculations",                      "data"),
    State({"type": "measure-select",  "index": ALL}, "options"),
    State({"type": "left-panel",      "index": ALL}, "id"),
    prevent_initial_call=True
)
def delete_calc(n_clicks, global_calcs, current_measure_options, panel_ids):
    if not any(n for n in n_clicks if n):
        raise dash.exceptions.PreventUpdate
    triggered = ctx.triggered_id
    if triggered is None:
        raise dash.exceptions.PreventUpdate
    name  = triggered["index"]
    calcs = dict(global_calcs or {})
    calcs.pop(name, None)
    try:
        cfg = load_config()
        cfg["global_calculations"] = calcs
        save_config(cfg)
    except Exception as e:
        print(f"Could not save calculations: {e}")
    updated_opts   = [[o for o in opts if o.get("value") != f"calc_{name}"]
                      for opts in current_measure_options]
    updated_panels = [build_table_panel(pid["index"], calcs) for pid in panel_ids]
    return calcs, updated_opts, updated_panels


@app.callback(
    Output("global-calculations",                    "data"),
    Output("calc-validation-msg",                    "children", allow_duplicate=True),
    Output({"type": "measure-select", "index": ALL}, "options"),
    Output({"type": "left-panel",     "index": ALL}, "children"),
    Output("calc-modal",                             "is_open",  allow_duplicate=True),
    Input("calc-save-btn", "n_clicks"),
    State("calc-name-input",    "value"),
    State("calc-type-select",   "value"),
    State("calc-agg-func",      "value"),
    State("calc-agg-field",     "value"),
    State("calc-formula-input", "value"),
    State("calc-lod-dim",       "value"),
    State("calc-lod-agg",       "value"),
    State("calc-lod-field",     "value"),
    State("global-calculations","data"),
    State({"type": "measure-select", "index": ALL}, "options"),
    State({"type": "left-panel",     "index": ALL}, "id"),
    State("editing-calc-name",  "data"),
    prevent_initial_call=True
)
def save_calculation(n_clicks, name, calc_type, agg_func, agg_field,
                     formula, lod_dim, lod_agg, lod_field,
                     existing_calcs, current_measure_options, panel_ids, editing_name):
    if not n_clicks:
        raise dash.exceptions.PreventUpdate
    if not name or not name.strip():
        return (existing_calcs,
                html.Span("⚠️ Please enter a name.", className="text-warning"),
                current_measure_options, [dash.no_update] * len(panel_ids), dash.no_update)
    name  = name.strip()
    calcs = dict(existing_calcs or {})
    if editing_name and editing_name in calcs:
        calcs.pop(editing_name)
        if editing_name != name:
            current_measure_options = [
                [o for o in opts if o.get("value") != f"calc_{editing_name}"]
                for opts in current_measure_options
            ]
    if name in calcs:
        return (existing_calcs,
                html.Span(f"⚠️ '{name}' already exists.", className="text-warning"),
                current_measure_options, [dash.no_update] * len(panel_ids), dash.no_update)
    if calc_type == "aggregate":
        defn    = {"type": "aggregate", "func": agg_func, "field": agg_field}
        display = f"{agg_func}({agg_field})"
    elif calc_type == "formula":
        if not formula or not formula.strip():
            return (existing_calcs,
                    html.Span("⚠️ Please enter a formula.", className="text-warning"),
                    current_measure_options, [dash.no_update] * len(panel_ids), dash.no_update)
        defn    = {"type": "formula", "formula": formula.strip()}
        display = formula.strip()
    elif calc_type == "fixed_lod":
        if not lod_dim or not lod_agg or not lod_field:
            return (existing_calcs,
                    html.Span("⚠️ Please fill in all LOD fields.", className="text-warning"),
                    current_measure_options, [dash.no_update] * len(panel_ids), dash.no_update)
        defn    = {"type": "fixed_lod", "dim": lod_dim, "agg": lod_agg, "field": lod_field}
        display = f"{lod_agg}({lod_field}) FIXED BY {lod_dim}"
    else:
        raise dash.exceptions.PreventUpdate

    calcs[name] = defn
    try:
        cfg = load_config()
        cfg["global_calculations"] = calcs
        save_config(cfg)
    except Exception as e:
        print(f"Could not save calculations: {e}")
    new_opt        = {"label": f"{name} — {display}", "value": f"calc_{name}"}
    updated_opts   = [opts + [new_opt] for opts in current_measure_options]
    updated_panels = [build_table_panel(pid["index"], calcs) for pid in panel_ids]
    return (calcs, html.Span(f"✅ '{name}' saved!", className="text-success"),
            updated_opts, updated_panels, False)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8050, debug=True)